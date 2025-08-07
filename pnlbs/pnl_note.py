import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from typing import Dict, List, Tuple, Any

class PnLGenerator:
    def __init__(self, json_file_path: str = "clean_financial_data_pnl.json"):
        """Initialize the P&L generator with JSON file path."""
        self.json_file_path = json_file_path
        self.financial_data = {}
        
    def load_financial_data(self) -> bool:
        """Load financial data from JSON file."""
        try:
            print(f" Loading financial data from: {self.json_file_path}")
            
            with open(self.json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Handle different JSON structures flexibly
            if "company_financial_data" in data:
                self.financial_data = data["company_financial_data"].get("other_data", {})
            elif "other_data" in data:
                self.financial_data = data["other_data"]
            else:
                # Assume the JSON structure matches your format directly
                self.financial_data = data
                
            print(f" Loaded data for {len(self.financial_data)} financial items")
            return True
            
        except FileNotFoundError:
            print(f" File not found: {self.json_file_path}")
            return False
        except json.JSONDecodeError as e:
            print(f" Invalid JSON format: {str(e)}")
            return False
        except Exception as e:
            print(f"Error loading data: {str(e)}")
            return False
    
    def extract_values(self, item_key: str) -> Tuple[float, float]:
        """Extract 2024 and 2023 values from financial data."""
        if item_key not in self.financial_data:
            print(f"  Warning: {item_key} not found in data")
            return 0.0, 0.0
        
        item_data = self.financial_data[item_key]
        total_2024 = 0.0
        total_2023 = 0.0
        
        # Handle different data structures
        if isinstance(item_data, dict):
            for category, values in item_data.items():
                if isinstance(values, list) and len(values) >= 2:
                    total_2024 += float(values[0] or 0)
                    total_2023 += float(values[1] or 0)
                elif isinstance(values, (int, float)):
                    total_2024 += float(values)
        elif isinstance(item_data, list) and len(item_data) >= 2:
            total_2024 = float(item_data[0] or 0)
            total_2023 = float(item_data[1] or 0)
        
        return total_2024, total_2023
    
    def get_revenue_data(self) -> Tuple[float, float]:
        """Extract revenue from operations data."""
        return self.extract_values("16. Revenue from Operations")
    
    def get_other_income_data(self) -> Tuple[float, float]:
        """Extract other income data."""
        return self.extract_values("17. Other income")
    
    def get_cost_materials_data(self) -> Tuple[float, float]:
        """Extract cost of materials consumed data."""
        # For cost of materials, we want the final calculated value
        item_key = "18. Cost of materials consumed"
        if item_key not in self.financial_data:
            print(f"  Warning: {item_key} not found in data")
            return 0.0, 0.0
            
        item_data = self.financial_data[item_key]
        
        # Look for the final "Cost of materials consumed" line
        if "Cost of materials consumed" in item_data:
            values = item_data["Cost of materials consumed"]
            if isinstance(values, list) and len(values) >= 2:
                return float(values[0] or 0), float(values[1] or 0)
        
        # Fallback: calculate from opening stock + purchases - closing stock
        opening_2024 = opening_2023 = 0.0
        purchases_2024 = purchases_2023 = 0.0
        closing_2024 = closing_2023 = 0.0
        
        if "Opening stock" in item_data:
            values = item_data["Opening stock"]
            if isinstance(values, list) and len(values) >= 2:
                opening_2024, opening_2023 = float(values[0] or 0), float(values[1] or 0)
        
        if "Add: Purchases" in item_data:
            values = item_data["Add: Purchases"]
            if isinstance(values, list) and len(values) >= 2:
                purchases_2024, purchases_2023 = float(values[0] or 0), float(values[1] or 0)
                
        if "Less: Closing stock" in item_data:
            values = item_data["Less: Closing stock"]
            if isinstance(values, list) and len(values) >= 2:
                closing_2024, closing_2023 = float(values[0] or 0), float(values[1] or 0)
        
        cost_2024 = opening_2024 + purchases_2024 - closing_2024
        cost_2023 = opening_2023 + purchases_2023 - closing_2023
        
        return cost_2024, cost_2023
    
    def get_employee_expense_data(self) -> Tuple[float, float]:
        """Extract employee benefit expense data."""
        return self.extract_values("19. Employee benefit expense")
    
    def get_other_expenses_data(self) -> Tuple[float, float]:
        """Extract other expenses data."""
        return self.extract_values("20. Other expenses")
    
    def get_depreciation_data(self) -> Tuple[float, float]:
        """Extract depreciation and amortisation data."""
        return self.extract_values("21. Depreciation and amortisation expense")
    
    def get_loss_on_sale_data(self) -> Tuple[float, float]:
        """Extract loss on sale of assets data."""
        return self.extract_values("22. Loss on sale of assets")
    
    def get_finance_costs_data(self) -> Tuple[float, float]:
        """Extract finance costs data."""
        return self.extract_values("23. Finance costs")
    
    def format_currency(self, value: float) -> str:
        """Format currency with commas."""
        if value == 0:
            return ""
        return f"{value:,.2f}"
    
    def generate_pnl_statement(self, output_file: str = "pnl_statement.xlsx") -> bool:
        """Generate comprehensive P&L statement Excel file."""
        
        if not self.financial_data:
            print(" No financial data loaded. Please load data first.")
            return False
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit and Loss Statement"

        # Define styles
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        bold_font = Font(bold=True, size=10)
        
        # Borders
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        top_bottom_border = Border(
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        # Alignments
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # Set column widths
        ws.column_dimensions["A"].width = 45  # Particulars
        ws.column_dimensions["B"].width = 8   # Notes
        ws.column_dimensions["C"].width = 15  # 2024
        ws.column_dimensions["D"].width = 15  # 2023

        row = 1

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = "Statement of Profit and Loss for the year ended March 31, 2024"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        ws["A1"].border = top_bottom_border
        row += 2

        # Currency annotation
        ws["C3"] = "In Lakhs"
        ws["C3"].font = normal_font
        ws["C3"].alignment = right_align
        row += 1

        # Column headers
        headers = ["", "Notes", "Year ended March 31, 2024", "Year ended March 31, 2023"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = top_bottom_border
            if col > 2:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        row += 1

        def add_data_row(description: str, note_ref: str, val_2024: float, val_2023: float, 
                        is_bold: bool = False, is_section_header: bool = False):
            """Add a data row with proper formatting."""
            nonlocal row
            
            # Description
            cell_a = ws.cell(row=row, column=1)
            cell_a.value = description
            cell_a.font = bold_font if (is_bold or is_section_header) else normal_font
            cell_a.alignment = left_align
            if not is_section_header:
                cell_a.border = thin_border
            
            # Note reference
            cell_b = ws.cell(row=row, column=2)
            cell_b.value = note_ref if note_ref else ""
            cell_b.font = normal_font
            cell_b.alignment = center_align
            if not is_section_header:
                cell_b.border = thin_border
            
            # 2024 value
            cell_c = ws.cell(row=row, column=3)
            cell_c.value = self.format_currency(val_2024)
            cell_c.font = bold_font if is_bold else normal_font
            cell_c.alignment = right_align
            if not is_section_header:
                cell_c.border = thin_border
            
            # 2023 value
            cell_d = ws.cell(row=row, column=4)
            cell_d.value = self.format_currency(val_2023)
            cell_d.font = bold_font if is_bold else normal_font
            cell_d.alignment = right_align
            if not is_section_header:
                cell_d.border = thin_border
            
            row += 1

        # Extract all financial data
        print("\n Extracting financial data...")
        
        # Income data
        revenue_2024, revenue_2023 = self.get_revenue_data()
        other_income_2024, other_income_2023 = self.get_other_income_data()
        
        # Expense data
        materials_2024, materials_2023 = self.get_cost_materials_data()
        employee_2024, employee_2023 = self.get_employee_expense_data()
        other_exp_2024, other_exp_2023 = self.get_other_expenses_data()
        depreciation_2024, depreciation_2023 = self.get_depreciation_data()
        loss_sale_2024, loss_sale_2023 = self.get_loss_on_sale_data()
        finance_2024, finance_2023 = self.get_finance_costs_data()

        # INCOME SECTION
        add_data_row("Income", "", 0, 0, is_section_header=True)
        add_data_row("Revenue from operations (net)", "16", revenue_2024, revenue_2023)
        add_data_row("Other income", "17", other_income_2024, other_income_2023)
        
        # Total revenue
        total_revenue_2024 = revenue_2024 + other_income_2024
        total_revenue_2023 = revenue_2023 + other_income_2023
        add_data_row("Total revenue (I)", "", total_revenue_2024, total_revenue_2023, is_bold=True)

        # EXPENSES SECTION
        add_data_row("Expenses", "", 0, 0, is_section_header=True)
        add_data_row("Cost of materials consumed", "18", materials_2024, materials_2023)
        add_data_row("Employee benefit expense", "19", employee_2024, employee_2023)
        add_data_row("Other expenses", "20", other_exp_2024, other_exp_2023)
        add_data_row("Depreciation and amortisation expense", "21", depreciation_2024, depreciation_2023)
        add_data_row("Loss on sale of assets & investments", "22", loss_sale_2024, loss_sale_2023)
        add_data_row("Finance costs", "23", finance_2024, finance_2023)
        
        # Total expenses
        total_expenses_2024 = materials_2024 + employee_2024 + other_exp_2024 + depreciation_2024 + loss_sale_2024 + finance_2024
        total_expenses_2023 = materials_2023 + employee_2023 + other_exp_2023 + depreciation_2023 + loss_sale_2023 + finance_2023
        add_data_row("Total Expenses (II)", "", total_expenses_2024, total_expenses_2023, is_bold=True)
        
        # Profit before tax
        profit_before_tax_2024 = total_revenue_2024 - total_expenses_2024
        profit_before_tax_2023 = total_revenue_2023 - total_expenses_2023
        add_data_row("Profit before Tax (I) - (II)", "", profit_before_tax_2024, profit_before_tax_2023, is_bold=True)
        
        # Tax Expense section (placeholders)
        add_data_row("IV. TAX EXPENSE", "", 0, 0, is_section_header=True)
        add_data_row("Current Tax", "", 0.0, 0.0)
        add_data_row("Deferred Tax Liability/(Asset)", "", 0.0, 0.0)
        add_data_row("Income Tax relating to Prior Year", "", 0.0, 0.0)
        add_data_row("MAT Credit (Entitlement)/Utilisation", "", 0.0, 0.0)
        add_data_row("Total Tax Expense (IV)", "", 0.0, 0.0, is_bold=True)
        
        # Profit after Tax (assuming no tax for now)
        profit_after_tax_2024 = profit_before_tax_2024
        profit_after_tax_2023 = profit_before_tax_2023
        add_data_row("Profit After Tax (III - IV)", "", profit_after_tax_2024, profit_after_tax_2023, is_bold=True)

        # Earnings per share section (placeholders)
        add_data_row("Earnings per share", "", 0, 0, is_section_header=True)
        add_data_row("Basic and diluted", "30", 0.0, 0.0)
        add_data_row("Nominal value", "", 10.0, 10.0)
        add_data_row("Weighted average number of equity shares", "30", 0.0, 0.0)

        # Footer
        row += 2
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "The accompanying notes are an integral part of the financial statements"
        ws[f"A{row}"].font = normal_font
        ws[f"A{row}"].alignment = left_align

        # Save the file
        try:
            wb.save(output_file)
            print(f"\n P&L Statement generated successfully: {output_file}")
            
            # Print financial summary
            self.print_financial_summary(total_revenue_2024, total_revenue_2023, 
                                       total_expenses_2024, total_expenses_2023,
                                       profit_before_tax_2024, profit_before_tax_2023,
                                       profit_after_tax_2024, profit_after_tax_2023)
            return True
            
        except PermissionError:
            print(f" Permission Error: Cannot save to {output_file}")
            fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "pnl_statement_fallback.xlsx")
            try:
                wb.save(fallback_file)
                print(f" P&L Statement saved to: {fallback_file}")
                return True
            except Exception as e:
                print(f" Failed to save: {str(e)}")
                return False
        except Exception as e:
            print(f" Error saving file: {str(e)}")
            return False
    
    def print_financial_summary(self, total_revenue_2024: float, total_revenue_2023: float,
                              total_expenses_2024: float, total_expenses_2023: float,
                              profit_before_tax_2024: float, profit_before_tax_2023: float,
                              profit_after_tax_2024: float, profit_after_tax_2023: float):
        """Print financial summary."""
        print("\n" + "="*60)
        print(" FINANCIAL SUMMARY")
        print("="*60)
        print(f"Total Revenue 2024:     Rs.{total_revenue_2024:>12,.2f} Lakhs")
        print(f"Total Revenue 2023:     Rs.{total_revenue_2023:>12,.2f} Lakhs")
        print(f"Total Expenses 2024:    Rs.{total_expenses_2024:>12,.2f} Lakhs")
        print(f"Total Expenses 2023:    Rs.{total_expenses_2023:>12,.2f} Lakhs")
        print(f"Profit Before Tax 2024: Rs.{profit_before_tax_2024:>12,.2f} Lakhs")
        print(f"Profit Before Tax 2023: Rs.{profit_before_tax_2023:>12,.2f} Lakhs")
        print(f"Profit After Tax 2024:  Rs.{profit_after_tax_2024:>12,.2f} Lakhs")
        print(f"Profit After Tax 2023:  Rs.{profit_after_tax_2023:>12,.2f} Lakhs")

        if total_revenue_2023 > 0:
            growth_rate = ((total_revenue_2024 - total_revenue_2023) / total_revenue_2023) * 100
            print(f"Revenue Growth Rate:    {growth_rate:>12.2f}%")


def main():
    """Main function to run the P&L generator."""
    print(" P&L STATEMENT GENERATOR FROM JSON")
    print("=" * 50)
    
    # You can specify different JSON file paths here
    json_files = [
        "clean_financial_data_pnl.json",
        "pnl_notes.json",
        
    ]
    
    # Try to find the JSON file
    json_file = None
    for file in json_files:
        if os.path.exists(file):
            json_file = file
            break
    
    if not json_file:
        json_file = input("Enter the path to your JSON file: ").strip()
    
    # Initialize generator
    generator = PnLGenerator(json_file)
    
    # Load data and generate P&L
    if generator.load_financial_data():
        output_path = "pnl_statement.xlsx"  # or wherever you save the file
        print(f"Output file: {output_path}")
        if generator.generate_pnl_statement("pnl_statement.xlsx"):
            print("\n P&L STATEMENT GENERATION COMPLETED SUCCESSFULLY!")
            print("Output file: pnl_statement.xlsx")  
        else:
            print("\n Failed to generate P&L statement")
    else:
        print("\n Failed to load financial data")


if __name__ == "__main__":
    main()