import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import re
import uuid

INPUT_FILE = "notestoALL/BLnotes2.xlsx"
OUTPUT_FILE = "notestoALL/outputbs_Sheet.xlsx"

def safe_float_conversion(value):
    """Safely convert value to float, handling various formats."""
    if value is None or str(value).strip() in ['#REF!', '-', '--', 'NA', '']:
        return 0.0
    str_val = str(value).strip().replace(',', '').replace('₹', '').replace('Rs.', '').replace(' ', '')
    is_negative = '(' in str_val and ')' in str_val
    str_val = str_val.replace('(', '').replace(')', '')
    try:
        return -float(str_val) if is_negative else float(str_val)
    except (ValueError, TypeError):
        return 0.0

def find_data_columns(ws):
    """Find the data columns for 2024 and 2023."""
    col_2024, col_2023 = None, None
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            if re.search(r'31\.?03\.?2024|45382|2024', cell_value, re.IGNORECASE):
                col_2024 = col_idx
            elif re.search(r'31\.?03\.?2023|45016|2023', cell_value, re.IGNORECASE):
                col_2023 = col_idx
    if not col_2024:
        col_2024 = 4  # Default to column D
    if not col_2023:
        col_2023 = 5  # Default to column E
    return col_2024, col_2023

def load_and_map_excel_notes(file_path=INPUT_FILE):
    """Load and parse balance sheet notes from Excel."""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        notes_data = {}
        
        print(f"Reading Excel file: {file_path}")
        print(f"Sheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        col_2024, col_2023 = find_data_columns(ws)
        print(f"Data columns - 2024: {col_2024}, 2023: {col_2023}")
        
        current_note = None
        current_note_data = None
        
        for row_idx in range(1, ws.max_row + 1):
            row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
            if not row_data or not row_data[0]:
                continue
            first_cell = str(row_data[0]).strip()
            
            note_match = re.match(r'^(\d{1,2})\.?\s*(.+)', first_cell)
            if note_match:
                note_num = note_match.group(1)
                if 2 <= int(note_num) <= 15:
                    if current_note and current_note_data:
                        notes_data[current_note] = current_note_data
                    current_note = note_num
                    category_name = note_match.group(2).strip()
                    current_note_data = {
                        "category_name": category_name,
                        "structure": [{
                            "category": category_name,
                            "subcategories": []
                        }],
                        "total_2024": 0.0,
                        "total_2023": 0.0,
                        "total_change": 0.0
                    }
                    print(f"\n📋 Processing Note {note_num}: {category_name}")
                    continue
            
            if current_note and current_note_data:
                # Refined skip keywords to avoid missing valid data
                skip_keywords = ['reconciliation', 'terms/ rights', 'details of', 'disclosure of',
                                'age wise', 'outstanding for', 'undisputed', 'disputed', 'sundry',
                                'gross carrying', 'accumulated depreciation', 'net carrying']
                if any(keyword in first_cell.lower() for keyword in skip_keywords):
                    continue
                
                value_2024 = safe_float_conversion(row_data[col_2024 - 1]) if col_2024 <= len(row_data) else 0.0
                value_2023 = safe_float_conversion(row_data[col_2023 - 1]) if col_2023 <= len(row_data) else 0.0
                
                if first_cell and (value_2024 != 0 or value_2023 != 0 or not re.match(r'^\d+\.?\s*', first_cell)):
                    subcategory = {
                        "label": first_cell,
                        "value": value_2024,
                        "previous_value": value_2023,
                        "change": value_2024 - value_2023,
                        "change_percent": ((value_2024 - value_2023) / value_2023 * 100) if value_2023 != 0 else 0
                    }
                    current_note_data["structure"][0]["subcategories"].append(subcategory)
                    current_note_data["total_2024"] += value_2024
                    current_note_data["total_2023"] += value_2023
                    print(f"  ✓ {first_cell[:40]:<40} | 2024: {value_2024:>10.2f} | 2023: {value_2023:>10.2f}")
        
        if current_note and current_note_data:
            current_note_data["total_change"] = current_note_data["total_2024"] - current_note_data["total_2023"]
            notes_data[current_note] = current_note_data
        
        json_file_path = "notestoALL/balance_sheet_notes.json"
        os.makedirs(os.path.dirname(json_file_path), exist_ok=True)
        with open(json_file_path, "w", encoding="utf-8") as f:
            json.dump(notes_data, f, indent=2, ensure_ascii=False)
        print(f"\n✅ JSON saved: {json_file_path}")
        print(f"📊 Total notes processed: {len(notes_data)}")
        
        print("\n" + "="*70)
        print("📈 NOTES SUMMARY")
        print("="*70)
        for note_num in sorted(notes_data.keys(), key=int):
            note = notes_data[note_num]
            items = len(note["structure"][0]["subcategories"])
            print(f"Note {note_num}: {note['category_name'][:35]:<35} | Items: {items:>3} | "
                  f"2024: {note['total_2024']:>10.2f} | 2023: {note['total_2023']:>10.2f}")
        
        return notes_data
    
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}

def generate_balance_sheet_report(notes_data):
    """Generate comprehensive Balance Sheet matching the template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    # Define styles
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    bold_font = Font(bold=True, size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
    top_bottom_border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    right_align = Alignment(horizontal="right", vertical="center")

    # Set column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    row = 1

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Balance Sheet as at March 31, 2024"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws["A1"].border = top_bottom_border
    row += 1

    # In Lakhs
    ws["C2"] = "In Lakhs"
    ws["C2"].font = normal_font
    ws["C2"].alignment = right_align
    row += 1

    # Headers
    headers = ["", "Notes", "March 31, 2024", "March 31, 2023"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = top_bottom_border
    row += 1

    def add_data_row(description, note_ref, val_2024, val_2023, indent=0, is_bold=False, is_section_header=False):
        """Add a data row with proper formatting."""
        nonlocal row
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = description
        cell_a.font = bold_font if (is_bold or is_section_header) else normal_font
        cell_a.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
        if not is_section_header:
            cell_a.border = thin_border
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = note_ref if note_ref else ""
        cell_b.font = normal_font
        cell_b.alignment = center_align
        if not is_section_header:
            cell_b.border = thin_border
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f"{val_2024:,.2f}" if val_2024 != 0 else ""
        cell_c.font = bold_font if is_bold else normal_font
        cell_c.alignment = right_align
        if not is_section_header:
            cell_c.border = thin_border
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f"{val_2023:,.2f}" if val_2023 != 0 else ""
        cell_d.font = bold_font if is_bold else normal_font
        cell_d.alignment = right_align
        if not is_section_header:
            cell_d.border = thin_border
        row += 1

    # Equity and Liabilities
    add_data_row("Equity and liabilities", "", 0, 0, is_section_header=True)
    
    # Shareholders' Funds
    add_data_row("Shareholders' funds", "", 0, 0, indent=1, is_section_header=True)
    
    # Share Capital (Note 2)
    share_capital_2024 = notes_data.get("2", {}).get("total_2024", 0.0)
    share_capital_2023 = notes_data.get("2", {}).get("total_2023", 0.0)
    if share_capital_2024 == 0 and share_capital_2023 == 0:
        print("⚠️ Warning: No data found for Note 2 (Share capital)")
    add_data_row("Share capital", "2", share_capital_2024, share_capital_2023, indent=2)
    
    # Reserves and Surplus (Note 3)
    reserves_2024 = notes_data.get("3", {}).get("total_2024", 0.0)
    reserves_2023 = notes_data.get("3", {}).get("total_2023", 0.0)
    if reserves_2024 == 0 and reserves_2023 == 0:
        print("⚠️ Warning: No data found for Note 3 (Reserves and surplus)")
    add_data_row("Reserves and surplus", "3", reserves_2024, reserves_2023, indent=2)
    
    # Total Shareholders' Funds
    total_equity_2024 = share_capital_2024 + reserves_2024
    total_equity_2023 = share_capital_2023 + reserves_2023
    add_data_row("", "", total_equity_2024, total_equity_2023, indent=1, is_bold=True)
    
    # Non-Current Liabilities
    add_data_row("Non-Current liabilities", "", 0, 0, indent=1, is_section_header=True)
    
    # Long Term Borrowings (Note 4)
    borrowings_2024 = notes_data.get("4", {}).get("total_2024", 0.0)
    borrowings_2023 = notes_data.get("4", {}).get("total_2023", 0.0)
    if borrowings_2024 == 0 and borrowings_2023 == 0:
        print("⚠️ Warning: No data found for Note 4 (Long term borrowings)")
    add_data_row("Long term borrowings", "4", borrowings_2024, borrowings_2023, indent=2)
    
    # Deferred Tax Liability (Note 5)
    deferred_tax_2024 = notes_data.get("5", {}).get("total_2024", 0.0)
    deferred_tax_2023 = notes_data.get("5", {}).get("total_2023", 0.0)
    if deferred_tax_2024 == 0 and deferred_tax_2023 == 0:
        print("⚠️ Warning: No data found for Note 5 (Deferred tax liability)")
    add_data_row("Deferred Tax Liability (Net)", "5", deferred_tax_2024, deferred_tax_2023, indent=2)
    
    # Total Non-Current Liabilities
    total_non_current_2024 = borrowings_2024 + deferred_tax_2024
    total_non_current_2023 = borrowings_2023 + deferred_tax_2023
    add_data_row("", "", total_non_current_2024, total_non_current_2023, indent=1, is_bold=True)
    
    # Current Liabilities
    add_data_row("Current liabilities", "", 0, 0, indent=1, is_section_header=True)
    
    # Trade Payables (Note 6)
    trade_payables_2024 = notes_data.get("6", {}).get("total_2024", 0.0)
    trade_payables_2023 = notes_data.get("6", {}).get("total_2023", 0.0)
    if trade_payables_2024 == 0 and trade_payables_2023 == 0:
        print("⚠️ Warning: No data found for Note 6 (Trade payables)")
    add_data_row("Trade payables", "6", trade_payables_2024, trade_payables_2023, indent=2)
    
    # Other Current Liabilities (Note 7)
    other_liabilities_2024 = notes_data.get("7", {}).get("total_2024", 0.0)
    other_liabilities_2023 = notes_data.get("7", {}).get("total_2023", 0.0)
    if other_liabilities_2024 == 0 and other_liabilities_2023 == 0:
        print("⚠️ Warning: No data found for Note 7 (Other current liabilities)")
    add_data_row("Other current liabilities", "7", other_liabilities_2024, other_liabilities_2023, indent=2)
    
    # Short Term Provisions (Note 8)
    provisions_2024 = notes_data.get("8", {}).get("total_2024", 0.0)
    provisions_2023 = notes_data.get("8", {}).get("total_2023", 0.0)
    if provisions_2024 == 0 and provisions_2023 == 0:
        print("⚠️ Warning: No data found for Note 8 (Short term provisions)")
    add_data_row("Short term provisions", "8", provisions_2024, provisions_2023, indent=2)
    
    # Total Current Liabilities
    total_current_liab_2024 = trade_payables_2024 + other_liabilities_2024 + provisions_2024
    total_current_liab_2023 = trade_payables_2023 + other_liabilities_2023 + provisions_2023
    add_data_row("", "", total_current_liab_2024, total_current_liab_2023, indent=1, is_bold=True)
    
    # Total Equity and Liabilities
    total_liabilities_2024 = total_equity_2024 + total_non_current_2024 + total_current_liab_2024
    total_liabilities_2023 = total_equity_2023 + total_non_current_2023 + total_current_liab_2023
    add_data_row("TOTAL", "", total_liabilities_2024, total_liabilities_2023, is_bold=True)
    
    # Assets
    add_data_row("Assets", "", 0, 0, is_section_header=True)
    
    # Non-Current Assets
    add_data_row("Non-current assets", "", 0, 0, indent=1, is_section_header=True)
    
    # Fixed Assets (Note 9)
    add_data_row("Fixed assets", "9", 0, 0, indent=2, is_section_header=True)
    
    # Tangible Assets
    tangible_2024 = sum(item["value"] for item in notes_data.get("9", {}).get("structure", [{}])[0].get("subcategories", []) 
                       if not any(x in item["label"].lower() for x in ["intangible", "software"]))
    tangible_2023 = sum(item["previous_value"] for item in notes_data.get("9", {}).get("structure", [{}])[0].get("subcategories", []) 
                       if not any(x in item["label"].lower() for x in ["intangible", "software"]))
    if tangible_2024 == 0 and tangible_2023 == 0:
        print("⚠️ Warning: No data found for Tangible assets (Note 9)")
    add_data_row("Tangible assets", "", tangible_2024, tangible_2023, indent=3)
    
    # Intangible Assets
    intangible_2024 = sum(item["value"] for item in notes_data.get("9", {}).get("structure", [{}])[0].get("subcategories", []) 
                         if any(x in item["label"].lower() for x in ["intangible", "software"]))
    intangible_2023 = sum(item["previous_value"] for item in notes_data.get("9", {}).get("structure", [{}])[0].get("subcategories", []) 
                         if any(x in item["label"].lower() for x in ["intangible", "software"]))
    if intangible_2024 == 0 and intangible_2023 == 0:
        print("⚠️ Warning: No data found for Intangible assets (Note 9)")
    add_data_row("Intangible assets", "", intangible_2024, intangible_2023, indent=3)
    
    # Capital Work in Progress
    capital_wip_2024, capital_wip_2023 = 0.0, 0.0
    print("⚠️ Warning: No data found for Capital Work in Progress")
    add_data_row("Capital Work in Progress", "", capital_wip_2024, capital_wip_2023, indent=3)
    
    # Long Term Loans and Advances (Note 10)
    long_term_loans_2024 = notes_data.get("10", {}).get("total_2024", 0.0)
    long_term_loans_2023 = notes_data.get("10", {}).get("total_2023", 0.0)
    if long_term_loans_2024 == 0 and long_term_loans_2023 == 0:
        print("⚠️ Warning: No data found for Note 10 (Long term loans and advances)")
    add_data_row("Long Term Loans and Advances", "10", long_term_loans_2024, long_term_loans_2023, indent=2)
    
    # Total Non-Current Assets
    total_non_current_assets_2024 = tangible_2024 + intangible_2024 + capital_wip_2024 + long_term_loans_2024
    total_non_current_assets_2023 = tangible_2023 + intangible_2023 + capital_wip_2023 + long_term_loans_2023
    add_data_row("", "", total_non_current_assets_2024, total_non_current_assets_2023, indent=1, is_bold=True)
    
    # Current Assets
    add_data_row("Current assets", "", 0, 0, indent=1, is_section_header=True)
    
    # Inventories (Note 11)
    inventories_2024 = notes_data.get("11", {}).get("total_2024", 0.0)
    inventories_2023 = notes_data.get("11", {}).get("total_2023", 0.0)
    if inventories_2024 == 0 and inventories_2023 == 0:
        print("⚠️ Warning: No data found for Note 11 (Inventories)")
    add_data_row("Inventories", "11", inventories_2024, inventories_2023, indent=2)
    
    # Trade Receivables (Note 12)
    trade_receivables_2024 = notes_data.get("12", {}).get("total_2024", 0.0)
    trade_receivables_2023 = notes_data.get("12", {}).get("total_2023", 0.0)
    if trade_receivables_2024 == 0 and trade_receivables_2023 == 0:
        print("⚠️ Warning: No data found for Note 12 (Trade receivables)")
    add_data_row("Trade receivables", "12", trade_receivables_2024, trade_receivables_2023, indent=2)
    
    # Cash and Bank Balances (Note 13)
    cash_balances_2024 = notes_data.get("13", {}).get("total_2024", 0.0)
    cash_balances_2023 = notes_data.get("13", {}).get("total_2023", 0.0)
    if cash_balances_2024 == 0 and cash_balances_2023 == 0:
        print("⚠️ Warning: No data found for Note 13 (Cash and bank balances)")
    add_data_row("Cash and bank balances", "13", cash_balances_2024, cash_balances_2023, indent=2)
    
    # Short Term Loans and Advances (Note 14)
    short_term_loans_2024 = notes_data.get("14", {}).get("total_2024", 0.0)
    short_term_loans_2023 = notes_data.get("14", {}).get("total_2023", 0.0)
    if short_term_loans_2024 == 0 and short_term_loans_2023 == 0:
        print("⚠️ Warning: No data found for Note 14 (Short term loans and advances)")
    add_data_row("Short-term loans and advances", "14", short_term_loans_2024, short_term_loans_2023, indent=2)
    
    # Other Current Assets (Note 15)
    other_assets_2024 = notes_data.get("15", {}).get("total_2024", 0.0)
    other_assets_2023 = notes_data.get("15", {}).get("total_2023", 0.0)
    if other_assets_2024 == 0 and other_assets_2023 == 0:
        print("⚠️ Warning: No data found for Note 15 (Other current assets)")
    add_data_row("Other current assets", "15", other_assets_2024, other_assets_2023, indent=2)
    
    # Total Current Assets
    total_current_assets_2024 = (inventories_2024 + trade_receivables_2024 + cash_balances_2024 + 
                                short_term_loans_2024 + other_assets_2024)
    total_current_assets_2023 = (inventories_2023 + trade_receivables_2023 + cash_balances_2023 + 
                                short_term_loans_2023 + other_assets_2023)
    add_data_row("", "", total_current_assets_2024, total_current_assets_2023, indent=1, is_bold=True)
    
    # Total Assets
    total_assets_2024 = total_non_current_assets_2024 + total_current_assets_2024
    total_assets_2023 = total_non_current_assets_2023 + total_current_assets_2023
    add_data_row("TOTAL", "", total_assets_2024, total_assets_2023, is_bold=True)
    
    # Footer Notes
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "The accompanying notes are an integral part of the financial statements"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "As per my report of even date. For and on behalf of the Board of Directors"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 2
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "For M/s Siva Parvathi & Associates"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "ICAI Firm registration number: 020872S"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Chartered Accountants"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 2
    
    ws[f"A{row}"] = "S. Siva Parvathi"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    ws[f"D{row}"] = "Director"
    ws[f"D{row}"].font = normal_font
    ws[f"D{row}"].alignment = right_align
    row += 1
    
    ws[f"A{row}"] = "Proprietor"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    ws[f"D{row}"] = "Director"
    ws[f"D{row}"].font = normal_font
    ws[f"D{row}"].alignment = right_align
    row += 1
    
    ws[f"A{row}"] = "Membership No.:"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws[f"A{row}"] = "UDIN: 24226087BKEECZ1200"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws[f"A{row}"] = "Place: Hyderabad"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws[f"A{row}"] = "Date: 04/09/2024"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align

    # Apply borders
    for r in range(1, row):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            if cell.value and not any(keyword in str(cell.value).lower() for keyword in ["balance sheet", "in lakhs", "notes"]):
                cell.border = thin_border

    # Save the file
    output_folder = "notestoALL"
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, "outputbs_Sheet.xlsx")
    try:
        wb.save(output_file)
        print(f"\n✅ Balance Sheet generated: {output_file}")
        
        # Print summary
        print("\n" + "="*60)
        print("📊 BALANCE SHEET SUMMARY")
        print("="*60)
        print(f"Total Equity 2024:         ₹{total_equity_2024:>12,.2f} Lakhs")
        print(f"Total Equity 2023:         ₹{total_equity_2023:>12,.2f} Lakhs")
        print(f"Total Non-Current Liab 2024: ₹{total_non_current_2024:>12,.2f} Lakhs")
        print(f"Total Non-Current Liab 2023: ₹{total_non_current_2023:>12,.2f} Lakhs")
        print(f"Total Current Liab 2024:   ₹{total_current_liab_2024:>12,.2f} Lakhs")
        print(f"Total Current Liab 2023:   ₹{total_current_liab_2023:>12,.2f} Lakhs")
        print(f"Total Assets 2024:         ₹{total_assets_2024:>12,.2f} Lakhs")
        print(f"Total Assets 2023:         ₹{total_assets_2023:>12,.2f} Lakhs")
    
    except PermissionError:
        print(f"❌ Permission Error: Cannot save to {output_file}")
        fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "balance_sheet_fallback.xlsx")
        try:
            wb.save(fallback_file)
            print(f"✅ Balance Sheet saved to: {fallback_file}")
        except Exception as e:
            print(f"❌ Failed to save: {str(e)}")
    except Exception as e:
        print(f"❌ Error saving file: {str(e)}")

def main():
    print("🚀 BALANCE SHEET GENERATOR")
    print("=" * 50)
    
    print("\n📋 STEP 1: Converting Excel to JSON")
    notes_data = load_and_map_excel_notes()
    
    if notes_data:
        print("\n📊 STEP 2: Generating Balance Sheet")
        generate_balance_sheet_report(notes_data)
        print("\n🎉 PROCESS COMPLETED!")
    else:
        print("\n❌ FAILED: No data found")

if __name__ == "__main__":
    main()