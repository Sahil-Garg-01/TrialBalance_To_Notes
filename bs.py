import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_note_data(note_number, folder="generated_notes"):
    """Load note data from JSON file in the specified folder."""
    file_path = os.path.join(folder, f"notes.json")
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Warning: Note {note_number} file not found at {file_path}")
        return None
    except json.JSONDecodeError:
        print(f"Warning: Invalid JSON in note {note_number}")
        return None

def extract_total_from_note(note_data, year="2024"):
    """Extract total value from note data for the specified year."""
    if not note_data or "structure" not in note_data:
        return 0.0
    
    # Look for total in categories
    for category in note_data["structure"]:
        if year == "2024" and "total" in category:
            try:
                value = float(category["total"])
                # Convert from paise to lakhs if value is very large
                if value > 100000:
                    value = value / 100000
                return value
            except (ValueError, TypeError):
                continue
        elif year == "2023" and "previous_total" in category:
            try:
                value = float(category["previous_total"])
                # Convert from paise to lakhs if value is very large
                if value > 100000:
                    value = value / 100000
                return value
            except (ValueError, TypeError):
                continue
    
    # If no total found, sum up subcategory values
    total = 0.0
    for category in note_data["structure"]:
        if "subcategories" in category:
            for subcat in category["subcategories"]:
                if year == "2024" and "value" in subcat:
                    try:
                        value = float(subcat["value"])
                        # Convert from paise to lakhs if value is very large
                        if value > 100000:
                            value = value / 100000
                        total += value
                    except (ValueError, TypeError):
                        continue
                elif year == "2023" and "previous_value" in subcat:
                    try:
                        value = float(subcat["previous_value"])
                        # Convert from paise to lakhs if value is very large
                        if value > 100000:
                            value = value / 100000
                        total += value
                    except (ValueError, TypeError):
                        continue
    
    return total

def extract_specific_value(note_data, key, year="2024"):
    """Extract specific value from note data based on key and year."""
    if not note_data or "structure" not in note_data:
        return 0.0
    
    # Search through all categories and subcategories
    for category in note_data["structure"]:
        if "subcategories" in category:
            for subcat in category["subcategories"]:
                label = subcat.get("label", "").lower().replace(" ", "").replace("-", "").replace("/", "").replace("&", "")
                if key.lower() in label:
                    if year == "2024" and "value" in subcat:
                        try:
                            value = float(subcat["value"])
                            # Convert from paise to lakhs if value is very large
                            if value > 100000:
                                value = value / 100000
                            return value
                        except (ValueError, TypeError):
                            continue
                    elif year == "2023" and "previous_value" in subcat:
                        try:
                            value = float(subcat["previous_value"])
                            # Convert from paise to lakhs if value is very large
                            if value > 100000:
                                value = value / 100000
                            return value
                        except (ValueError, TypeError):
                            continue
    
    return 0.0

def format_currency(value):
    """Format currency value for display."""
    if isinstance(value, (int, float)) and value != 0:
        return f"{value:,.2f}"
    return "0.00"

def generate_balance_sheet_report():
    """Generate Balance Sheet report in Excel format using data from generated_notes folder."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    # Define styles
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    right_align = Alignment(horizontal="right")

    # Set column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Header
    ws["A1"] = "Balance Sheet as at March 31, 2024"
    ws["A1"].font = bold_font
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = center_align

    # Units
    ws["A2"] = "In Lakhs"
    ws.merge_cells("A2:D2")
    ws["A2"].alignment = right_align

    # Table headers
    headers = ["", "Notes", "March 31, 2024", "March 31, 2023"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align if col > 1 else left_align

    # Load all required notes (2-15 for Balance Sheet)
    notes_data = {}
    for note_num in range(2, 16):  # Notes 2-15
        notes_data[str(note_num)] = load_note_data(str(note_num))

    # Extract values from notes
    # Equity and Liabilities
    share_capital_2024 = extract_total_from_note(notes_data.get("2"), "2024")
    share_capital_2023 = extract_total_from_note(notes_data.get("2"), "2023")
    reserves_surplus_2024 = extract_total_from_note(notes_data.get("3"), "2024")
    reserves_surplus_2023 = extract_total_from_note(notes_data.get("3"), "2023")
    
    # Non-current liabilities
    long_term_borrowings_2024 = extract_total_from_note(notes_data.get("4"), "2024")
    long_term_borrowings_2023 = extract_total_from_note(notes_data.get("4"), "2023")
    deferred_tax_liability_2024 = extract_total_from_note(notes_data.get("5"), "2024")
    deferred_tax_liability_2023 = extract_total_from_note(notes_data.get("5"), "2023")
    
    # Current liabilities
    trade_payables_2024 = extract_total_from_note(notes_data.get("6"), "2024")
    trade_payables_2023 = extract_total_from_note(notes_data.get("6"), "2023")
    other_current_liabilities_2024 = extract_total_from_note(notes_data.get("7"), "2024")
    other_current_liabilities_2023 = extract_total_from_note(notes_data.get("7"), "2023")
    short_term_provisions_2024 = extract_total_from_note(notes_data.get("8"), "2024")
    short_term_provisions_2023 = extract_total_from_note(notes_data.get("8"), "2023")

    # Assets
    # Non-current assets
    fixed_assets_2024 = extract_total_from_note(notes_data.get("9"), "2024")
    fixed_assets_2023 = extract_total_from_note(notes_data.get("9"), "2023")
    long_term_loans_advances_2024 = extract_total_from_note(notes_data.get("10"), "2024")
    long_term_loans_advances_2023 = extract_total_from_note(notes_data.get("10"), "2023")
    
    # Current assets
    inventories_2024 = extract_total_from_note(notes_data.get("11"), "2024")
    inventories_2023 = extract_total_from_note(notes_data.get("11"), "2023")
    trade_receivables_2024 = extract_total_from_note(notes_data.get("12"), "2024")
    trade_receivables_2023 = extract_total_from_note(notes_data.get("12"), "2023")
    cash_bank_balances_2024 = extract_total_from_note(notes_data.get("13"), "2024")
    cash_bank_balances_2023 = extract_total_from_note(notes_data.get("13"), "2023")
    short_term_loans_advances_2024 = extract_total_from_note(notes_data.get("14"), "2024")
    short_term_loans_advances_2023 = extract_total_from_note(notes_data.get("14"), "2023")
    other_current_assets_2024 = extract_total_from_note(notes_data.get("15"), "2024")
    other_current_assets_2023 = extract_total_from_note(notes_data.get("15"), "2023")

    # Calculate totals
    shareholders_funds_2024 = share_capital_2024 + reserves_surplus_2024
    shareholders_funds_2023 = share_capital_2023 + reserves_surplus_2023
    
    non_current_liabilities_2024 = long_term_borrowings_2024 + deferred_tax_liability_2024
    non_current_liabilities_2023 = long_term_borrowings_2023 + deferred_tax_liability_2023
    
    current_liabilities_2024 = trade_payables_2024 + other_current_liabilities_2024 + short_term_provisions_2024
    current_liabilities_2023 = trade_payables_2023 + other_current_liabilities_2023 + short_term_provisions_2023
    
    total_equity_liabilities_2024 = shareholders_funds_2024 + non_current_liabilities_2024 + current_liabilities_2024
    total_equity_liabilities_2023 = shareholders_funds_2023 + non_current_liabilities_2023 + current_liabilities_2023
    
    non_current_assets_2024 = fixed_assets_2024 + long_term_loans_advances_2024
    non_current_assets_2023 = fixed_assets_2023 + long_term_loans_advances_2023
    
    current_assets_2024 = (inventories_2024 + trade_receivables_2024 + cash_bank_balances_2024 + 
                          short_term_loans_advances_2024 + other_current_assets_2024)
    current_assets_2023 = (inventories_2023 + trade_receivables_2023 + cash_bank_balances_2023 + 
                          short_term_loans_advances_2023 + other_current_assets_2023)
    
    total_assets_2024 = non_current_assets_2024 + current_assets_2024
    total_assets_2023 = non_current_assets_2023 + current_assets_2023

   


    # Balance Sheet line items
    line_items = [
        # Equity and Liabilities
        {"label": "Equity and liabilities", "note": "", "value_2024": "", "value_2023": "", "is_header": True},
        {"label": "Shareholders' funds", "note": "", "value_2024": "", "value_2023": "", "is_subheader": True},
        {"label": "Share capital", "note": "2", "value_2024": share_capital_2024, "value_2023": share_capital_2023},
        {"label": "Reserves and surplus", "note": "3", "value_2024": reserves_surplus_2024, "value_2023": reserves_surplus_2023},
        {"label": "", "note": "", "value_2024": shareholders_funds_2024, "value_2023": shareholders_funds_2023, "is_total": True},
        
        {"label": "Non-Current liabilities", "note": "", "value_2024": "", "value_2023": "", "is_subheader": True},
        {"label": "Long term borrowings", "note": "4", "value_2024": long_term_borrowings_2024, "value_2023": long_term_borrowings_2023},
        {"label": "Deferred Tax Liability (Net)", "note": "5", "value_2024": deferred_tax_liability_2024, "value_2023": deferred_tax_liability_2023},
        {"label": "", "note": "", "value_2024": non_current_liabilities_2024, "value_2023": non_current_liabilities_2023, "is_total": True},
        
        {"label": "Current liabilities", "note": "", "value_2024": "", "value_2023": "", "is_subheader": True},
        {"label": "Trade payables", "note": "6", "value_2024": trade_payables_2024, "value_2023": trade_payables_2023},
        {"label": "Other current liabilities", "note": "7", "value_2024": other_current_liabilities_2024, "value_2023": other_current_liabilities_2023},
        {"label": "Short term provisions", "note": "8", "value_2024": short_term_provisions_2024, "value_2023": short_term_provisions_2023},
        {"label": "", "note": "", "value_2024": current_liabilities_2024, "value_2023": current_liabilities_2023, "is_total": True},
        
        {"label": "TOTAL", "note": "", "value_2024": total_equity_liabilities_2024, "value_2023": total_equity_liabilities_2023, "is_grand_total": True},
        
        # Assets
        {"label": "", "note": "", "value_2024": "", "value_2023": "", "is_spacer": True},
        {"label": "Assets", "note": "", "value_2024": "", "value_2023": "", "is_header": True},
        {"label": "Non-current assets", "note": "", "value_2024": "", "value_2023": "", "is_subheader": True},
        {"label": "Fixed assets", "note": "9", "value_2024": fixed_assets_2024, "value_2023": fixed_assets_2023},
        {"label": "Long Term Loans and Advances", "note": "10", "value_2024": long_term_loans_advances_2024, "value_2023": long_term_loans_advances_2023},
        {"label": "", "note": "", "value_2024": non_current_assets_2024, "value_2023": non_current_assets_2023, "is_total": True},
        
        {"label": "Current assets", "note": "", "value_2024": "", "value_2023": "", "is_subheader": True},
        {"label": "Inventories", "note": "11", "value_2024": inventories_2024, "value_2023": inventories_2023},
        {"label": "Trade receivables", "note": "12", "value_2024": trade_receivables_2024, "value_2023": trade_receivables_2023},
        {"label": "Cash and bank balances", "note": "13", "value_2024": cash_bank_balances_2024, "value_2023": cash_bank_balances_2023},
        {"label": "Short-term loans and advances", "note": "14", "value_2024": short_term_loans_advances_2024, "value_2023": short_term_loans_advances_2023},
        {"label": "Other current assets", "note": "15", "value_2024": other_current_assets_2024, "value_2023": other_current_assets_2023},
        {"label": "", "note": "", "value_2024": current_assets_2024, "value_2023": current_assets_2023, "is_total": True},
        
        {"label": "TOTAL", "note": "", "value_2024": total_assets_2024, "value_2023": total_assets_2023, "is_grand_total": True}
    ]

    # Write line items to Excel
    row = 5
    for item in line_items:
        # Skip spacer rows
        if item.get("is_spacer"):
            row += 1
            continue
            
        ws.cell(row=row, column=1).value = item["label"]
        ws.cell(row=row, column=2).value = item["note"]
        
        # Format values
        if item["value_2024"] == "":
            ws.cell(row=row, column=3).value = ""
        else:
            ws.cell(row=row, column=3).value = format_currency(item["value_2024"])
            
        if item["value_2023"] == "":
            ws.cell(row=row, column=4).value = ""
        else:
            ws.cell(row=row, column=4).value = format_currency(item["value_2023"])
        
        # Apply formatting
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align if col > 1 else left_align
            
            # Apply special formatting based on item type
            if item.get("is_header") or item.get("is_grand_total"):
                ws.cell(row=row, column=col).font = bold_font
            elif item.get("is_subheader"):
                ws.cell(row=row, column=col).font = bold_font
            elif item.get("is_total"):
                ws.cell(row=row, column=col).font = bold_font
        
        row += 1

    # Add footer note
    row += 1
    ws.cell(row=row, column=1).value = "The accompanying notes are an integral part of the financial statements."
    ws.cell(row=row, column=1).alignment = left_align
    row += 1
    ws.cell(row=row, column=1).value = "As per my report of even date."
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=4).value = "For and on behalf of the Board of Directors"
    ws.cell(row=row, column=4).alignment = center_align

    # Save Excel file with error handling
    output_file = "balance_sheet_report.xlsx"
    try:
        wb.save(output_file)
        print(f"Balance Sheet report generated successfully and saved to {output_file}")
        
    except PermissionError:
        print(f"PermissionError: Unable to save to {output_file}. Trying alternative location...")
        fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "balance_sheet_report_fallback.xlsx")
        try:
            wb.save(fallback_file)
            print(f"Balance Sheet report saved to alternative location: {fallback_file}")
        except Exception as e:
            print(f"Failed to save Balance Sheet report: {str(e)}")
    except Exception as e:
        print(f"Error saving Balance Sheet report: {str(e)}")

if __name__ == "__main__":
    generate_balance_sheet_report()