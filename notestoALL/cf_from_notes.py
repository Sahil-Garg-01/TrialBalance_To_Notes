import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import re
from datetime import datetime

INPUT_FILE = "notestoALL/NotesToCF.xlsx"
OUTPUT_FILE = "outputcf_sheet.xlsx"

def safe_float_conversion(value):
    """Safely convert value to float, handling various formats."""
    if value is None:
        return 0.0
    str_val = str(value).strip()
    if not str_val or str_val in ['-', '--', 'NA', '#REF!', '']:
        return 0.0
    str_val = str_val.replace(',', '').replace('₹', '').replace('Rs.', '').replace(' ', '')
    is_negative = '(' in str_val and ')' in str_val
    str_val = str_val.replace('(', '').replace(')', '')
    try:
        result = float(str_val)
        return -result if is_negative else result
    except (ValueError, TypeError):
        return 0.0

def find_data_columns(ws):
    """Find the data columns for 2024 and 2023."""
    for row_idx in range(1, 11):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            if re.search(r'31\.?03\.?2024|2024', cell_value, re.IGNORECASE):
                col_2024 = col_idx
            elif re.search(r'31\.?03\.?2023|2023', cell_value, re.IGNORECASE):
                col_2023 = col_idx
    if 'col_2024' not in locals():
        col_2024 = ws.max_column - 1 if ws.max_column > 2 else 2
    if 'col_2023' not in locals():
        col_2023 = ws.max_column if ws.max_column > 1 else 3
    return col_2024, col_2023

def load_and_map_excel_notes(file_path=INPUT_FILE):
    """Load and parse cash flow notes from Excel."""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        notes_data = {}
        
        print(f"Reading Excel file: {file_path}")
        print(f"Sheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        col_2024, col_2023 = find_data_columns(ws)
        print(f"Data columns - 2024: {col_2024}, 2023: {col_2023}")
        
        current_note = None
        current_note_data = None
        
        for row_idx in range(1, ws.max_row + 1):
            row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
            if not row_data or not row_data[0]:
                continue
            first_cell = str(row_data[0]).strip()
            
            note_match = re.match(r'^(\d{2})\.?\s*(.+)', first_cell)
            if note_match:
                note_num = note_match.group(1)
                if int(note_num) >= 24:  # Start from relevant notes (e.g., EPS, Depreciation)
                    if current_note and current_note_data:
                        notes_data[current_note] = current_note_data
                    current_note = note_num
                    category_name = note_match.group(2).strip()
                    current_note_data = {
                        "category_name": category_name,
                        "structure": [{
                            "category": category_name,
                            "subcategories": []
                        }],
                        "total_2024": 0.0,
                        "total_2023": 0.0,
                        "total_change": 0.0
                    }
                    print(f"\n📋 Processing Note {note_num}: {category_name}")
                    continue
            
            if current_note and current_note_data:
                skip_keywords = ['in lakhs', 'march 31', 'particulars', 'year ended', 'amount', 
                               'total', 'subtotal', '2024', '2023', 'as per', 'pursuant']
                if (len(first_cell) <= 2 or 
                    any(keyword in first_cell.lower() for keyword in skip_keywords)):
                    continue
                
                value_2024 = safe_float_conversion(row_data[col_2024 - 1]) if col_2024 <= len(row_data) else 0.0
                value_2023 = safe_float_conversion(row_data[col_2023 - 1]) if col_2023 <= len(row_data) else 0.0
                
                if (first_cell and len(first_cell.strip()) > 2 and 
                    (value_2024 != 0.0 or value_2023 != 0.0 or any(c.isalpha() for c in first_cell))):
                    subcategory = {
                        "label": first_cell,
                        "value": value_2024,
                        "previous_value": value_2023,
                        "change": value_2024 - value_2023,
                        "change_percent": ((value_2024 - value_2023) / value_2023 * 100) if value_2023 != 0 else 0
                    }
                    current_note_data["structure"][0]["subcategories"].append(subcategory)
                    current_note_data["total_2024"] += value_2024
                    current_note_data["total_2023"] += value_2023
                    print(f"  ✓ {first_cell[:40]:<40} | 2024: {value_2024:>10.2f} | 2023: {value_2023:>10.2f}")
        
        if current_note and current_note_data:
            current_note_data["total_change"] = current_note_data["total_2024"] - current_note_data["total_2023"]
            notes_data[current_note] = current_note_data
        
        
        for note_num in sorted(notes_data.keys(), key=int):
            note = notes_data[note_num]
            items = len(note["structure"][0]["subcategories"])
            print(f"Note {note_num}: {note['category_name'][:35]:<35} | Items: {items:>3} | "
                  f"2024: {note['total_2024']:>10.2f} | 2023: {note['total_2023']:>10.2f}")
        
        return notes_data
    
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}

def generate_cash_flow_report(notes_data):
    """Generate comprehensive Cash Flow Statement matching the template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Statement"

    # Define styles
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    bold_font = Font(bold=True, size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
    top_bottom_border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    right_align = Alignment(horizontal="right", vertical="center")
    fill_grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Set column widths
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    row = 1

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "Statement of Cash flows for the year ended March 31, 2024"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws["A1"].border = top_bottom_border
    row += 1

    # In Lakhs
    ws.merge_cells("A2:C2")
    ws["A2"] = "In Lakhs"
    ws["A2"].font = normal_font
    ws["A2"].alignment = center_align
    row += 1

    # Headers
    headers = ["Particulars", "March 31, 2024", "March 31, 2023"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = top_bottom_border
        cell.fill = fill_grey
    row += 1

    def add_data_row(description, val_2024, val_2023, indent=0, is_bold=False, is_section_header=False):
        """Add a data row with proper formatting."""
        nonlocal row
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = description
        cell_a.font = bold_font if (is_bold or is_section_header) else normal_font
        cell_a.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
        if not is_section_header:
            cell_a.border = thin_border
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"{val_2024:,.2f}" if val_2024 != 0 else ""
        cell_b.font = bold_font if is_bold else normal_font
        cell_b.alignment = right_align
        if not is_section_header:
            cell_b.border = thin_border
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f"{val_2023:,.2f}" if val_2023 != 0 else ""
        cell_c.font = bold_font if is_bold else normal_font
        cell_c.alignment = right_align
        if not is_section_header:
            cell_c.border = thin_border
        row += 1

    # Operating Activities
    add_data_row("Cash flow from operating activities", 0, 0, is_section_header=True)
    
    # Profit before tax (using Note 28 for EPS-related data as a proxy)
    profit_before_tax_2024 = notes_data.get("28", {}).get("total_2024", 0.0)
    profit_before_tax_2023 = notes_data.get("28", {}).get("total_2023", 0.0)
    if profit_before_tax_2024 == 0 and profit_before_tax_2023 == 0:
        print("⚠️ Warning: No data found for Profit before tax (Note 28)")
    add_data_row("Profit before taxation", profit_before_tax_2024, profit_before_tax_2023, indent=1)
    
    add_data_row("Adjustment for :", 0, 0, indent=1)
    
    # Depreciation (Note 30)
    depreciation_2024 = notes_data.get("30", {}).get("total_2024", 0.0)
    depreciation_2023 = notes_data.get("30", {}).get("total_2023", 0.0)
    if depreciation_2024 == 0 and depreciation_2023 == 0:
        print("⚠️ Warning: No data found for Note 30 (Depreciation)")
    add_data_row("Add: Depreciation and Amortisation Expense", depreciation_2024, depreciation_2023, indent=2)
    
    # Interest income (not found in notes, placeholder)
    interest_income_2024 = 0.0
    interest_income_2023 = 0.0
    print("⚠️ Warning: No data found for Interest income")
    add_data_row("Less: Interest income", interest_income_2024, interest_income_2023, indent=2)
    
    # Operating profit
    op_profit_2024 = profit_before_tax_2024 + depreciation_2024 + interest_income_2024
    op_profit_2023 = profit_before_tax_2023 + depreciation_2023 + interest_income_2023
    add_data_row("Operating profit before working capital changes", op_profit_2024, op_profit_2023, is_bold=True)
    
    # Working Capital Movements (using Note 31 for ratios as proxies)
    add_data_row("Movements in working capital:", 0, 0, indent=1, is_section_header=True)
    wc_items = {
        "trade_receivables": "(Increase)/Decrease in Trade Receivables",
        "other_current_assets": "(Increase)/Decrease in Other Current Assets",
        "trade_payables": "Increase/(Decrease) in Trade Payables"
    }
    wc_2024, wc_2023 = 0.0, 0.0
    for key, label in wc_items.items():
        val_2024 = notes_data.get("31", {}).get("structure", [{}])[0].get("subcategories", [])
        val_2024 = next((item["value"] for item in val_2024 if key in item["label"].lower()), 0.0)
        val_2023 = notes_data.get("31", {}).get("structure", [{}])[0].get("subcategories", [])
        val_2023 = next((item["previous_value"] for item in val_2023 if key in item["label"].lower()), 0.0)
        if val_2024 == 0 and val_2023 == 0:
            print(f"⚠️ Warning: No data found for {label} (Note 31)")
        add_data_row(label, val_2024, val_2023, indent=2)
        wc_2024 += val_2024
        wc_2023 += val_2023
    
    # Placeholder for other working capital items
    other_wc_items = [
        "(Increase)/Decrease in Inventories",
        "(Increase)/Decrease in Short Term Loans & Advances",
        "(Increase)/Decrease in Capital Work in Progress",
        "(Increase)/Decrease in Long Term Loans & Advances",
        "Increase/(Decrease) in Short Term Provisions",
        "Increase/(Decrease) in Other Current Liabilities"
    ]
    for label in other_wc_items:
        print(f"⚠️ Warning: No data found for {label}")
        add_data_row(label, 0.0, 0.0, indent=2)
    
    cash_ops_2024 = op_profit_2024 + wc_2024
    cash_ops_2023 = op_profit_2023 + wc_2023
    add_data_row("Cash used in operations", cash_ops_2024, cash_ops_2023, indent=1, is_bold=True)
    
    # Taxes paid (placeholder)
    taxes_2024, taxes_2023 = 0.0, 0.0
    print("⚠️ Warning: No data found for Direct taxes paid")
    add_data_row("Less: Direct taxes paid (net of refunds)", taxes_2024, taxes_2023, indent=1)
    
    net_cash_ops_2024 = cash_ops_2024 - taxes_2024
    net_cash_ops_2023 = cash_ops_2023 - taxes_2023
    add_data_row("Net cash flow used in operating activities (A)", net_cash_ops_2024, net_cash_ops_2023, is_bold=True)
    
    # Investing Activities
    add_data_row("Cash flows from investing activities", 0, 0, is_section_header=True)
    purchase_assets_2024, purchase_assets_2023 = 0.0, 0.0
    sale_assets_2024, sale_assets_2023 = 0.0, 0.0
    inv_interest_2024, inv_interest_2023 = 0.0, 0.0
    print("⚠️ Warning: No data found for Purchase of Assets")
    print("⚠️ Warning: No data found for Sale of Assets")
    print("⚠️ Warning: No data found for Interest income (Investing)")
    add_data_row("Purchase of Assets", purchase_assets_2024, purchase_assets_2023, indent=1)
    add_data_row("Sale of Assets", sale_assets_2024, sale_assets_2023, indent=1)
    add_data_row("Interest income", inv_interest_2024, inv_interest_2023, indent=1)
    net_cash_inv_2024 = purchase_assets_2024 + sale_assets_2024 + inv_interest_2024
    net_cash_inv_2023 = purchase_assets_2023 + sale_assets_2023 + inv_interest_2023
    add_data_row("Net cash flow used in investing activities (B)", net_cash_inv_2024, net_cash_inv_2023, is_bold=True)
    
    # Financing Activities
    add_data_row("Cash flows from financing activities", 0, 0, is_section_header=True)
    dividend_2024, dividend_2023 = 0.0, 0.0
    borrowings_2024 = next((item["value"] for item in notes_data.get("31", {}).get("structure", [{}])[0].get("subcategories", []) 
                           if "debt-equity ratio" in item["label"].lower()), 0.0)
    borrowings_2023 = next((item["previous_value"] for item in notes_data.get("31", {}).get("structure", [{}])[0].get("subcategories", []) 
                           if "debt-equity ratio" in item["label"].lower()), 0.0)
    if borrowings_2024 == 0 and borrowings_2023 == 0:
        print("⚠️ Warning: No data found for Long Term Borrowings (Note 31)")
    print("⚠️ Warning: No data found for Dividend paid")
    add_data_row("Dividend paid", dividend_2024, dividend_2023, indent=1)
    add_data_row("Long Term Borrowings", borrowings_2024, borrowings_2023, indent=1)
    net_cash_fin_2024 = dividend_2024 + borrowings_2024
    net_cash_fin_2023 = dividend_2023 + borrowings_2023
    add_data_row("Net cash generated from financing activities (C)", net_cash_fin_2024, net_cash_fin_2023, is_bold=True)
    
    # Net Cash Flow
    net_cash_2024 = net_cash_ops_2024 + net_cash_inv_2024 + net_cash_fin_2024
    net_cash_2023 = net_cash_ops_2023 + net_cash_inv_2023 + net_cash_fin_2023
    add_data_row("Net increase in cash and cash equivalents (A+B+C)", net_cash_2024, net_cash_2023, indent=2, is_bold=True)
    
    # Cash balances (placeholders)
    cash_begin_2024, cash_begin_2023 = 0.0, 0.0
    print("⚠️ Warning: No data found for Cash and cash equivalents at the beginning")
    add_data_row("Cash and cash equivalents at the beginning of the year", cash_begin_2024, cash_begin_2023)
    cash_end_2024 = cash_begin_2024 + net_cash_2024
    cash_end_2023 = cash_begin_2023 + net_cash_2023
    add_data_row("Cash and cash equivalents at the end of the year", cash_end_2024, cash_end_2023, is_bold=True)
    
    # Cash Equivalents Components
    add_data_row("Components of cash and cash equivalents", 0, 0, is_section_header=True)
    cash_hand_2024, cash_hand_2023 = 0.0, 0.0
    bank_current_2024, bank_current_2023 = 0.0, 0.0
    bank_fixed_2024, bank_fixed_2023 = 0.0, 0.0
    print("⚠️ Warning: No data found for Cash on hand")
    print("⚠️ Warning: No data found for Bank Current Accounts")
    print("⚠️ Warning: No data found for Bank Fixed Deposits")
    add_data_row("Cash on hand", cash_hand_2024, cash_hand_2023, indent=1)
    add_data_row("With banks in Current Accounts", bank_current_2024, bank_current_2023, indent=1)
    add_data_row("With banks in Fixed Deposits", bank_fixed_2024, bank_fixed_2023, indent=1)
    add_data_row("Total cash and cash equivalents (Refer note 13)", cash_end_2024, cash_end_2023, indent=1, is_bold=True)
    
    # Notes
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Notes:"
    ws[f"A{row}"].font = bold_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    notes = [
        "1. The Cash Flow statement is prepared under 'indirect method' as set out in the Indian Accounting Standard - 7 on Cash Flow Statements. Cash and cash equivalents in the Cash Flow Statement comprise cash at bank and in hand and deposits with bank.",
        "2. Previous year's figures have been regrouped, wherever necessary.",
        "3. The accompanying notes form an integral part of the Financial Statements"
    ]
    for note in notes:
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = note
        ws[f"A{row}"].font = normal_font
        ws[f"A{row}"].alignment = left_align
        row += 1
    
    # Signatures
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "As per my report of even date. For and on behalf of the Board of Directors"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = center_align
    row += 2
    ws[f"A{row}"] = "For M/s Siva Parvathi & Associates"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    ws[f"C{row}"] = "Director"
    ws[f"C{row}"].font = normal_font
    ws[f"C{row}"].alignment = right_align
    row += 1
    ws[f"A{row}"] = "ICAI Firm registration number: 020872S"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    ws[f"C{row}"] = "Director"
    ws[f"C{row}"].font = normal_font
    ws[f"C{row}"].alignment = right_align
    row += 1
    ws[f"A{row}"] = "Chartered Accountants"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 2
    ws[f"A{row}"] = "S. Siva Parvathi"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    ws[f"A{row}"] = "Proprietor"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    ws[f"A{row}"] = "Membership No.: "
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    ws[f"A{row}"] = "UDIN: 24226087BKEECZ1200"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    ws[f"A{row}"] = "Place: Hyderabad"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    ws[f"A{row}"] = "Date: 04/09/2024"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align

    # Apply borders
    for r in range(1, row):
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            if cell.value and not any(keyword in str(cell.value).lower() for keyword in ["statement of", "in lakhs", "particulars"]):
                cell.border = thin_border

    # Save the file
    output_folder = "notestoALL"
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, "outputcf_sheet.xlsx")
    try:
        wb.save(output_file)
        print(f"\n✅ Cash Flow Statement generated: {output_file}")

        # Print summary
        print("\n" + "="*60)
        print("📊 CASH FLOW SUMMARY")
        print("="*60)
        print(f"Net Cash from Operating 2024: ₹{net_cash_ops_2024:>12,.2f} Lakhs")
        print(f"Net Cash from Operating 2023: ₹{net_cash_ops_2023:>12,.2f} Lakhs")
        print(f"Net Cash from Investing 2024: ₹{net_cash_inv_2024:>12,.2f} Lakhs")
        print(f"Net Cash from Investing 2023: ₹{net_cash_inv_2023:>12,.2f} Lakhs")
        print(f"Net Cash from Financing 2024: ₹{net_cash_fin_2024:>12,.2f} Lakhs")
        print(f"Net Cash from Financing 2023: ₹{net_cash_fin_2023:>12,.2f} Lakhs")
        print(f"Net Increase in Cash 2024:    ₹{net_cash_2024:>12,.2f} Lakhs")
        print(f"Net Increase in Cash 2023:    ₹{net_cash_2023:>12,.2f} Lakhs")
    
    except PermissionError:
        print(f"❌ Permission Error: Cannot save to {OUTPUT_FILE}")
        fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "cash_flow_statement_fallback.xlsx")
        try:
            wb.save(fallback_file)
            print(f"✅ Cash Flow Statement saved to: {fallback_file}")
        except Exception as e:
            print(f"❌ Failed to save: {str(e)}")
    except Exception as e:
        print(f"❌ Error saving file: {str(e)}")

def main():
    print("🚀 CASH FLOW STATEMENT GENERATOR")
    print("=" * 50)
    
    print("\n📋 STEP 1: Converting Excel to JSON")
    notes_data = load_and_map_excel_notes()
    
    if notes_data:
        print("\n📊 STEP 2: Generating Cash Flow Statement")
        generate_cash_flow_report(notes_data)
        print("\n🎉 PROCESS COMPLETED!")
    else:
        print("\n❌ FAILED: No data found")

if __name__ == "__main__":
    main()