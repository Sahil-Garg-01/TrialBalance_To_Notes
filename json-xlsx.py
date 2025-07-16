import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def create_output_folder(folder_path):
    """Create output folder if it doesn't exist"""
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Created folder: {folder_path}")

def read_json_file(file_path):
    """Read and parse JSON file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        print(f"Successfully read JSON file: {file_path}")
        return data
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON format in '{file_path}': {e}")
        return None
    except Exception as e:
        print(f"Error reading file '{file_path}': {e}")
        return None

def create_financial_table_sheet(workbook, sheet_name, note_data):
    """Create a properly formatted financial table sheet"""
    ws = workbook.create_sheet(title=sheet_name)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Add Note Title
    note_title = note_data.get('full_title', note_data.get('note_title', 'Note'))
    ws.cell(row=current_row, column=1, value=note_title)
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    current_row += 2
    
    # Process table_data if available
    if 'table_data' in note_data and note_data['table_data']:
        table_data = note_data['table_data']
        
        # Create DataFrame from table_data
        df = pd.DataFrame(table_data)
        
        # Add table headers
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=current_row, column=col_num, value=column_name.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Add table data
        for _, row in df.iterrows():
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = thin_border
                
                # Right align numeric columns (except first column)
                if col_num > 1:
                    cell.alignment = right_alignment
                
                # Bold formatting for total rows and headers
                if isinstance(value, str) and ('**' in value or 'Total' in value or 'Particulars' in value):
                    cell.font = bold_font
                    # Remove markdown formatting
                    cell.value = value.replace('**', '')
        
            current_row += 1
        
        current_row += 1
    
    # Add breakdown information if available
    if 'breakdown' in note_data and note_data['breakdown']:
        ws.cell(row=current_row, column=1, value="Breakdown Details:")
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        
        # Headers for breakdown
        ws.cell(row=current_row, column=1, value="Description")
        ws.cell(row=current_row, column=2, value="Amount")
        ws.cell(row=current_row, column=3, value="Amount (Lakhs)")
        
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Add breakdown data
        for key, value in note_data['breakdown'].items():
            if isinstance(value, dict):
                desc = value.get('description', key)
                amount = value.get('amount', 0)
                amount_lakhs = value.get('amount_lakhs', 0)
                
                ws.cell(row=current_row, column=1, value=desc).border = thin_border
                ws.cell(row=current_row, column=2, value=amount).border = thin_border
                ws.cell(row=current_row, column=3, value=amount_lakhs).border = thin_border
                
                # Right align numeric columns
                ws.cell(row=current_row, column=2).alignment = right_alignment
                ws.cell(row=current_row, column=3).alignment = right_alignment
                
                current_row += 1
        
        current_row += 1
    
    # Add matched accounts if available
    if 'matched_accounts' in note_data and note_data['matched_accounts']:
        ws.cell(row=current_row, column=1, value="Account-wise Breakdown:")
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        
        # Headers for matched accounts
        headers = ["Account", "Amount", "Amount (Lakhs)", "Group"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Add matched accounts data
        for account in note_data['matched_accounts']:
            ws.cell(row=current_row, column=1, value=account.get('account', '')).border = thin_border
            ws.cell(row=current_row, column=2, value=account.get('amount', 0)).border = thin_border
            ws.cell(row=current_row, column=3, value=account.get('amount_lakhs', 0)).border = thin_border
            ws.cell(row=current_row, column=4, value=account.get('group', '')).border = thin_border
            
            # Right align numeric columns
            ws.cell(row=current_row, column=2).alignment = right_alignment
            ws.cell(row=current_row, column=3).alignment = right_alignment
            
            current_row += 1
        
        current_row += 1
    
    # Add summary information
    if 'total_amount' in note_data:
        ws.cell(row=current_row, column=1, value="Summary:")
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Total Amount:")
        ws.cell(row=current_row, column=2, value=note_data.get('total_amount', 0))
        ws.cell(row=current_row, column=2).alignment = right_alignment
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Total Amount (Lakhs):")
        ws.cell(row=current_row, column=2, value=note_data.get('total_amount_lakhs', 0))
        ws.cell(row=current_row, column=2).alignment = right_alignment
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Matched Accounts Count:")
        ws.cell(row=current_row, column=2, value=note_data.get('matched_accounts_count', 0))
        ws.cell(row=current_row, column=2).alignment = right_alignment
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 60)  # Cap at 60 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return ws

def convert_json_to_excel(input_file, output_file):
    """Main function to convert JSON to Excel"""
    # Read JSON data
    json_data = read_json_file(input_file)
    if json_data is None:
        return False
    
    # Create workbook
    workbook = Workbook()
    
    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # Handle different JSON structures
    if 'notes' in json_data:
        # If JSON has 'notes' key, process each note
        notes_data = json_data['notes']
        for note in notes_data:
            note_title = note.get('full_title', note.get('note_title', f"Note {note.get('note_number', '')}"))
            # Clean sheet name
            clean_sheet_name = str(note_title).replace('/', '_').replace('\\', '_').replace('*', '_')
            clean_sheet_name = clean_sheet_name.replace('?', '_').replace('[', '_').replace(']', '_')
            clean_sheet_name = clean_sheet_name[:31]  # Excel sheet name limit
            
            print(f"Processing: {clean_sheet_name}")
            create_financial_table_sheet(workbook, clean_sheet_name, note)
    else:
        # Process each top-level key as a separate sheet
        for note_key, note_data in json_data.items():
            print(f"Processing: {note_key}")
            
            # Clean sheet name
            clean_sheet_name = str(note_key).replace('/', '_').replace('\\', '_').replace('*', '_')
            clean_sheet_name = clean_sheet_name.replace('?', '_').replace('[', '_').replace(']', '_')
            clean_sheet_name = clean_sheet_name[:31]  # Excel sheet name limit
            
            if isinstance(note_data, dict):
                create_financial_table_sheet(workbook, clean_sheet_name, note_data)
            else:
                # Handle non-dict top-level values
                simple_data = {"value": note_data}
                create_financial_table_sheet(workbook, clean_sheet_name, simple_data)
    
    # Save the workbook
    try:
        workbook.save(output_file)
        print(f"Successfully saved Excel file: {output_file}")
        return True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False

def main():
    """Main execution function"""
    # Define file paths
    input_file = "output2/notes_output.json"
    output_folder = "output3"
    output_file = os.path.join(output_folder, "final_notes_output.xlsx")
    
    # Create output folder
    create_output_folder(output_folder)
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found.")
        print("Please ensure the file exists in the correct location.")
        return
    
    # Convert JSON to Excel
    success = convert_json_to_excel(input_file, output_file)
    
    if success:
        print("\n" + "="*50)
        print("CONVERSION COMPLETED SUCCESSFULLY!")
        print("="*50)
        print(f"Input file: {input_file}")
        print(f"Output file: {output_file}")
        print("\nThe Excel file has been created with:")
        print("- Each note as a separate sheet")
        print("- Proper financial table formatting")
        print("- Table data displayed in tabular format")
        print("- Breakdown and account details included")
        print("- Professional styling and formatting")
    else:
        print("\n" + "="*50)
        print("CONVERSION FAILED!")
        print("="*50)
        print("Please check the error messages above.")

if __name__ == "__main__":
    main()