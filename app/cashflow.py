import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_data(file_path):
    """Load data from a JSON file with error handling."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Warning: File not found at {file_path}. Please ensure the file exists in the specified folder.")
        return {}
    except json.JSONDecodeError:
        print(f"Warning: Invalid JSON in {file_path}")
        return {}

def load_note_data(note_number, folder=None):
    import os
    import json
    if folder is None:
        folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), "generated_notes")
    file_path = os.path.join(folder, "notes.json")
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            notes = data.get("notes", [])
            for note in notes:
                n_num = note.get("note_number") or note.get("metadata", {}).get("note_number")
                if str(n_num) == str(note_number):
                    return note
            print(f"Warning: Note {note_number} not found in {file_path}")
            return None
    except FileNotFoundError:
        print(f"Warning: Note {note_number} file not found at {file_path}")
        return None
    except json.JSONDecodeError:
        print(f"Warning: Invalid JSON in note {note_number}")
        return None

def load_trail_balance(folder="output1"):
    """Load parsed trial balance data from JSON file in the specified folder."""
    return load_data(os.path.join(folder, "parsed_trail_balance.json"))

def extract_value_from_notes(note_data, key, year="2024"):
    """Extract specific value from note data based on key and year."""
    if not note_data or "structure" not in note_data:
        return None
    key = key.lower().replace(" ", "").replace("-", "").replace("/", "").replace("&", "")
    for category in note_data["structure"]:
        if "subcategories" in category:
            for subcat in category["subcategories"]:
                label = subcat.get("label", "").lower().replace(" ", "").replace("-", "").replace("/", "").replace("&", "")
                if key in label:
                    if year == "2024" and "value" in subcat:
                        try:
                            value = float(subcat["value"])
                            if value > 100000:  # Convert from paise to lakhs
                                value = value / 100000
                            return value
                        except (ValueError, TypeError):
                            continue
                    elif year == "2023" and "previous_value" in subcat:
                        try:
                            value = float(subcat["previous_value"])
                            if value > 100000:  # Convert from paise to lakhs
                                value = value / 100000
                            return value
                        except (ValueError, TypeError):
                            continue
    return None

def extract_value_from_tb(tb_data, account_name, year="2024"):
    """Extract balance from trial balance for a specific account and year."""
    if not tb_data or "accounts" not in tb_data.get(year, {}):
        return None
    for account in tb_data[year]["accounts"]:
        if account.get("name", "").lower() == account_name.lower():
            return float(account.get("balance", 0)) / 100000 if float(account.get("balance", 0)) > 100000 else float(account.get("balance", 0))
    return None

def calculate_movement(tb_2024, tb_2023, account_name):
    """Calculate movement (2024 - 2023) from trial balance data."""
    val_2024 = extract_value_from_tb({"2024": tb_2024}, account_name, "2024")
    val_2023 = extract_value_from_tb({"2023": tb_2023}, account_name, "2023")
    if val_2024 is not None and val_2023 is not None:
        return val_2024 - val_2023
    return None

def format_currency(value):
    """Format currency value for display, handling None."""
    if value is None:
        return "-"
    if isinstance(value, (int, float)) and value != 0:
        return f"{value:,.2f}"
    return "-"

def generate_cashflow_report():
    """Generate Cash Flow Statement report in Excel format using notes and trial balance data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Statement"

    # Define styles
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    right_align = Alignment(horizontal="right")

    # Set column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    # Header
    ws["A1"] = "Statement of Cash Flows for the year ended March 31, 2024"
    ws["A1"].font = bold_font
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = center_align

    # Units
    ws["A2"] = "In Lakhs"
    ws.merge_cells("A2:C2")
    ws["A2"].alignment = right_align

    # Table headers
    headers = ["Particulars", "March 31, 2024", "March 31, 2023"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align if col > 1 else left_align

    # Load data
    note_4 = load_note_data("4")  # Cash Flow Statement
    note_13 = load_note_data("13")  # Cash and Cash Equivalents
    note_6 = load_note_data("6")  # Trade Payables
    note_7 = load_note_data("7")  # Other Current Liabilities
    note_8 = load_note_data("8")  # Provisions
    note_9 = load_note_data("9")  # Fixed Assets
    tb_data = load_trail_balance()
    tb_2024 = tb_data.get("2024", {})
    tb_2023 = tb_data.get("2023", {})

    # Extract or calculate values
    values = {
        "profit_before_tax": {"2024": extract_value_from_notes(note_4, "profit before taxation", "2024") or extract_value_from_tb(tb_2024, "Profit before Taxation"),
                              "2023": extract_value_from_notes(note_4, "profit before taxation", "2023") or extract_value_from_tb(tb_2023, "Profit before Taxation")},
        "depreciation_amortization": {"2024": extract_value_from_notes(note_4, "depreciation and amortisation expense", "2024") or extract_value_from_tb(tb_2024, "Depreciation and Amortisation Expense"),
                                      "2023": extract_value_from_notes(note_4, "depreciation and amortisation expense", "2023") or extract_value_from_tb(tb_2023, "Depreciation and Amortisation Expense")},
        "interest_income": {"2024": extract_value_from_notes(note_4, "interest income", "2024") or extract_value_from_tb(tb_2024, "Interest Income"),
                            "2023": extract_value_from_notes(note_4, "interest income", "2023") or extract_value_from_tb(tb_2023, "Interest Income")},
        "trade_receivables": {"2024": calculate_movement(tb_2024, tb_2023, "Trade Receivables") or extract_value_from_notes(note_4, "increase/decrease in trade receivables", "2024"),
                              "2023": calculate_movement(tb_2023, tb_2023, "Trade Receivables") or extract_value_from_notes(note_4, "increase/decrease in trade receivables", "2023")},
        "inventories": {"2024": calculate_movement(tb_2024, tb_2023, "Inventories") or extract_value_from_notes(note_4, "increase/decrease in inventories", "2024"),
                        "2023": calculate_movement(tb_2023, tb_2023, "Inventories") or extract_value_from_notes(note_4, "increase/decrease in inventories", "2023")},
        "other_current_assets": {"2024": calculate_movement(tb_2024, tb_2023, "Other Current Assets") or extract_value_from_notes(note_4, "increase/decrease in other current assets", "2024"),
                                 "2023": calculate_movement(tb_2023, tb_2023, "Other Current Assets") or extract_value_from_notes(note_4, "increase/decrease in other current assets", "2023")},
        "short_term_loans": {"2024": calculate_movement(tb_2024, tb_2023, "Short Term Loans & Advances") or extract_value_from_notes(note_4, "increase/decrease in short term loans & advances", "2024"),
                             "2023": calculate_movement(tb_2023, tb_2023, "Short Term Loans & Advances") or extract_value_from_notes(note_4, "increase/decrease in short term loans & advances", "2023")},
        "capital_work_progress": {"2024": calculate_movement(tb_2024, tb_2023, "Capital Work in Progress") or extract_value_from_notes(note_4, "increase/decrease in capital work in progress", "2024"),
                                  "2023": calculate_movement(tb_2023, tb_2023, "Capital Work in Progress") or extract_value_from_notes(note_4, "increase/decrease in capital work in progress", "2023")},
        "long_term_loans": {"2024": calculate_movement(tb_2024, tb_2023, "Long Term Loans & Advances") or extract_value_from_notes(note_4, "increase/decrease in long term loans & advances", "2024"),
                            "2023": calculate_movement(tb_2023, tb_2023, "Long Term Loans & Advances") or extract_value_from_notes(note_4, "increase/decrease in long term loans & advances", "2023")},
        "short_term_provisions": {"2024": calculate_movement(tb_2024, tb_2023, "Short Term Provisions") or extract_value_from_notes(note_8, "short term provisions", "2024"),
                                  "2023": calculate_movement(tb_2023, tb_2023, "Short Term Provisions") or extract_value_from_notes(note_8, "short term provisions", "2023")},
        "trade_payables": {"2024": calculate_movement(tb_2024, tb_2023, "Trade Payables") or extract_value_from_notes(note_6, "trade payables", "2024"),
                           "2023": calculate_movement(tb_2023, tb_2023, "Trade Payables") or extract_value_from_notes(note_6, "trade payables", "2023")},
        "other_current_liabilities": {"2024": calculate_movement(tb_2024, tb_2023, "Other Current Liabilities") or extract_value_from_notes(note_7, "other current liabilities", "2024"),
                                      "2023": calculate_movement(tb_2023, tb_2023, "Other Current Liabilities") or extract_value_from_notes(note_7, "other current liabilities", "2023")},
        "purchase_assets": {"2024": extract_value_from_notes(note_9, "purchase of assets", "2024") or calculate_movement(tb_2024, tb_2023, "Fixed Assets"),
                            "2023": extract_value_from_notes(note_9, "purchase of assets", "2023") or calculate_movement(tb_2023, tb_2023, "Fixed Assets")},
        "sale_assets": {"2024": extract_value_from_notes(note_9, "sale of assets", "2024"),
                        "2023": extract_value_from_notes(note_9, "sale of assets", "2023")},
        "dividend_paid": {"2024": extract_value_from_notes(note_4, "dividend paid", "2024") or extract_value_from_tb(tb_2024, "Dividend Paid"),
                          "2023": extract_value_from_notes(note_4, "dividend paid", "2023") or extract_value_from_tb(tb_2023, "Dividend Paid")},
        "long_term_borrowings": {"2024": calculate_movement(tb_2024, tb_2023, "Long Term Borrowings") or extract_value_from_notes(note_4, "long term borrowings", "2024"),
                                 "2023": calculate_movement(tb_2023, tb_2023, "Long Term Borrowings") or extract_value_from_notes(note_4, "long term borrowings", "2023")},
        "cash_equivalents_opening": {"2024": extract_value_from_notes(note_13, "cash and cash equivalents at the beginning", "2024") or extract_value_from_tb(tb_2023, "Cash and Cash Equivalents"),
                                     "2023": extract_value_from_notes(note_13, "cash and cash equivalents at the beginning", "2023") or extract_value_from_tb(tb_2023, "Cash and Cash Equivalents")},
        "cash_equivalents_closing": {"2024": extract_value_from_notes(note_13, "cash and cash equivalents at the end", "2024") or extract_value_from_tb(tb_2024, "Cash and Cash Equivalents"),
                                     "2023": extract_value_from_notes(note_13, "cash and cash equivalents at the end", "2023") or extract_value_from_tb(tb_2023, "Cash and Cash Equivalents")},
        "cash_on_hand": {"2024": extract_value_from_notes(note_13, "cash on hand", "2024") or extract_value_from_tb(tb_2024, "Cash on Hand"),
                         "2023": extract_value_from_notes(note_13, "cash on hand", "2023") or extract_value_from_tb(tb_2023, "Cash on Hand")},
        "bank_current_accounts": {"2024": extract_value_from_notes(note_13, "with banks in current accounts", "2024") or extract_value_from_tb(tb_2024, "Bank Current Accounts"),
                                  "2023": extract_value_from_notes(note_13, "with banks in current accounts", "2023") or extract_value_from_tb(tb_2023, "Bank Current Accounts")},
        "bank_fixed_deposits": {"2024": extract_value_from_notes(note_13, "with banks in fixed deposits", "2024") or extract_value_from_tb(tb_2024, "Bank Fixed Deposits"),
                                "2023": extract_value_from_notes(note_13, "with banks in fixed deposits", "2023") or extract_value_from_tb(tb_2023, "Bank Fixed Deposits")},
    }

    # Calculate derived values
    operating_profit_2024 = (values["profit_before_tax"]["2024"] or 0) + (values["depreciation_amortization"]["2024"] or 0) - (values["interest_income"]["2024"] or 0)
    operating_profit_2023 = (values["profit_before_tax"]["2023"] or 0) + (values["depreciation_amortization"]["2023"] or 0) - (values["interest_income"]["2023"] or 0)
    cash_used_operations_2024 = operating_profit_2024 + sum(v["2024"] for k, v in values.items() if k.startswith("trade_") or k.startswith("inventories") or k.startswith("other_current_") or k.startswith("short_term_") or k.startswith("capital_work_") or k.startswith("long_term_") or k.startswith("short_term_provisions") or k.startswith("trade_payables") or k.startswith("other_current_liabilities") if v["2024"] is not None)
    cash_used_operations_2023 = operating_profit_2023 + sum(v["2023"] for k, v in values.items() if k.startswith("trade_") or k.startswith("inventories") or k.startswith("other_current_") or k.startswith("short_term_") or k.startswith("capital_work_") or k.startswith("long_term_") or k.startswith("short_term_provisions") or k.startswith("trade_payables") or k.startswith("other_current_liabilities") if v["2023"] is not None)
    direct_taxes_paid_2024 = extract_value_from_notes(note_8, "provision for taxation", "2024")
    direct_taxes_paid_2023 = extract_value_from_notes(note_8, "provision for taxation", "2023")
    net_cash_operating_2024 = cash_used_operations_2024 - (direct_taxes_paid_2024 or 0)
    net_cash_operating_2023 = cash_used_operations_2023 - (direct_taxes_paid_2023 or 0)
    net_cash_investing_2024 = -(values["purchase_assets"]["2024"] or 0) + (values["sale_assets"]["2024"] or 0) + (values["interest_income"]["2024"] or 0)
    net_cash_investing_2023 = -(values["purchase_assets"]["2023"] or 0) + (values["sale_assets"]["2023"] or 0) + (values["interest_income"]["2023"] or 0)
    net_cash_financing_2024 = (values["long_term_borrowings"]["2024"] or 0) - (values["dividend_paid"]["2024"] or 0)
    net_cash_financing_2023 = (values["long_term_borrowings"]["2023"] or 0) - (values["dividend_paid"]["2023"] or 0)
    total_cash_flow_2024 = net_cash_operating_2024 + net_cash_investing_2024 + net_cash_financing_2024
    total_cash_flow_2023 = net_cash_operating_2023 + net_cash_investing_2023 + net_cash_financing_2023

    # Debug prints
    print(f"Debug - Total Cash Flow 2024: {format_currency(total_cash_flow_2024)}")
    print(f"Debug - Total Cash Flow 2023: {format_currency(total_cash_flow_2023)}")
    print(f"Debug - Net Cash from Operating Activities 2024: {format_currency(net_cash_operating_2024)}")
    print(f"Debug - Net Cash from Operating Activities 2023: {format_currency(net_cash_operating_2023)}")

    # Cash Flow Statement line items
    line_items = [
        {"label": "Cash flow from operating activities", "value_2024": "", "value_2023": "", "is_header": True},
        {"label": "Profit before taxation", "value_2024": values["profit_before_tax"]["2024"], "value_2023": values["profit_before_tax"]["2023"]},
        {"label": "Adjustment for:", "value_2024": "", "value_2023": "", "is_subheader": True},
        {"label": "Add: Depreciation and Amortisation Expense", "value_2024": values["depreciation_amortization"]["2024"], "value_2023": values["depreciation_amortization"]["2023"]},
        {"label": "Less: Interest income", "value_2024": -values["interest_income"]["2024"] if values["interest_income"]["2024"] is not None else None, "value_2023": -values["interest_income"]["2023"] if values["interest_income"]["2023"] is not None else None},
        {"label": "Operating profit before working capital changes", "value_2024": operating_profit_2024, "value_2023": operating_profit_2023, "is_total": True},
        {"label": "Movements in working capital:", "value_2024": "", "value_2023": "", "is_subheader": True},
        {"label": "(Increase)/Decrease in Trade Receivables", "value_2024": values["trade_receivables"]["2024"], "value_2023": values["trade_receivables"]["2023"]},
        {"label": "(Increase)/Decrease in Inventories", "value_2024": values["inventories"]["2024"], "value_2023": values["inventories"]["2023"]},
        {"label": "(Increase)/Decrease in Other Current Assets", "value_2024": values["other_current_assets"]["2024"], "value_2023": values["other_current_assets"]["2023"]},
        {"label": "(Increase)/Decrease in Short Term Loans & Advances", "value_2024": values["short_term_loans"]["2024"], "value_2023": values["short_term_loans"]["2023"]},
        {"label": "(Increase)/Decrease in Capital Work in Progress", "value_2024": values["capital_work_progress"]["2024"], "value_2023": values["capital_work_progress"]["2023"]},
        {"label": "(Increase)/Decrease in Long Term Loans & Advances", "value_2024": values["long_term_loans"]["2024"], "value_2023": values["long_term_loans"]["2023"]},
        {"label": "Increase/(Decrease) in Short Term Provisions", "value_2024": values["short_term_provisions"]["2024"], "value_2023": values["short_term_provisions"]["2023"]},
        {"label": "Increase/(Decrease) in Trade Payables", "value_2024": values["trade_payables"]["2024"], "value_2023": values["trade_payables"]["2023"]},
        {"label": "Increase/(Decrease) in Other Current Liabilities", "value_2024": values["other_current_liabilities"]["2024"], "value_2023": values["other_current_liabilities"]["2023"]},
        {"label": "Cash used in operations", "value_2024": cash_used_operations_2024, "value_2023": cash_used_operations_2023, "is_total": True},
        {"label": "Less: Direct taxes paid (net of refunds)", "value_2024": direct_taxes_paid_2024, "value_2023": direct_taxes_paid_2023},
        {"label": "Net cash flow used in operating activities", "value_2024": net_cash_operating_2024, "value_2023": net_cash_operating_2023, "is_total": True},
        {"label": "Cash flows from investing activities", "value_2024": "", "value_2023": "", "is_header": True},
        {"label": "Purchase of Assets", "value_2024": -values["purchase_assets"]["2024"] if values["purchase_assets"]["2024"] is not None else None, "value_2023": -values["purchase_assets"]["2023"] if values["purchase_assets"]["2023"] is not None else None},
        {"label": "Sale of Assets", "value_2024": values["sale_assets"]["2024"], "value_2023": values["sale_assets"]["2023"]},
        {"label": "Interest income", "value_2024": values["interest_income"]["2024"], "value_2023": values["interest_income"]["2023"]},
        {"label": "Net cash flow used in investing activities", "value_2024": net_cash_investing_2024, "value_2023": net_cash_investing_2023, "is_total": True},
        {"label": "Cash flows from financing activities", "value_2024": "", "value_2023": "", "is_header": True},
        {"label": "Dividend paid", "value_2024": -values["dividend_paid"]["2024"] if values["dividend_paid"]["2024"] is not None else None, "value_2023": -values["dividend_paid"]["2023"] if values["dividend_paid"]["2023"] is not None else None},
        {"label": "Long Term Borrowings", "value_2024": values["long_term_borrowings"]["2024"], "value_2023": values["long_term_borrowings"]["2023"]},
        {"label": "Net cash generated from financing activities", "value_2024": net_cash_financing_2024, "value_2023": net_cash_financing_2023, "is_total": True},
        {"label": "Net increase/(decrease) in cash and cash equivalents", "value_2024": total_cash_flow_2024, "value_2023": total_cash_flow_2023, "is_total": True},
        {"label": "Cash and cash equivalents at the beginning of the year", "value_2024": values["cash_equivalents_opening"]["2024"], "value_2023": values["cash_equivalents_opening"]["2023"]},
        {"label": "Cash and cash equivalents at the end of the year", "value_2024": values["cash_equivalents_closing"]["2024"], "value_2023": values["cash_equivalents_closing"]["2023"], "is_grand_total": True},
        {"label": "Components of cash and cash equivalents", "value_2024": "", "value_2023": "", "is_header": True},
        {"label": "Cash on hand", "value_2024": values["cash_on_hand"]["2024"], "value_2023": values["cash_on_hand"]["2023"]},
        {"label": "With banks in Current Accounts", "value_2024": values["bank_current_accounts"]["2024"], "value_2023": values["bank_current_accounts"]["2023"]},
        {"label": "With banks in Fixed Deposits", "value_2024": values["bank_fixed_deposits"]["2024"], "value_2023": values["bank_fixed_deposits"]["2023"]},
        {"label": "Total cash and cash equivalents (Refer note 13)", "value_2024": values["cash_equivalents_closing"]["2024"], "value_2023": values["cash_equivalents_closing"]["2023"], "is_total": True},
    ]

    # Write line items to Excel
    row = 5
    for item in line_items:
        ws.cell(row=row, column=1).value = item["label"]
        ws.cell(row=row, column=2).value = format_currency(item["value_2024"])
        ws.cell(row=row, column=3).value = format_currency(item["value_2023"])
        
        # Apply formatting
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align if col > 1 else left_align
            if item.get("is_header") or item.get("is_grand_total"):
                ws.cell(row=row, column=col).font = bold_font
            elif item.get("is_subheader"):
                ws.cell(row=row, column=col).font = bold_font
            elif item.get("is_total"):
                ws.cell(row=row, column=col).font = bold_font
        row += 1

    # Add notes and footer
    row += 1
    ws.cell(row=row, column=1).value = "Notes:"
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=1).alignment = left_align
    row += 1
    ws.cell(row=row, column=1).value = "1. The Cash Flow statement is prepared under 'indirect method' as set out in the Indian Accounting Standard - 7 on Cash Flow Statements. Cash and cash equivalents in the Cash Flow Statement comprise cash at bank and in hand and deposits with bank."
    ws.cell(row=row, column=1).alignment = left_align
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    ws.cell(row=row, column=1).value = "2. Previous year's figures have been regrouped, wherever necessary."
    ws.cell(row=row, column=1).alignment = left_align
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    ws.cell(row=row, column=1).value = "3. The accompanying notes form an integral part of the Financial Statements"
    ws.cell(row=row, column=1).alignment = left_align
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    ws.cell(row=row, column=1).value = "As per my report of even date."
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=3).value = "For and on behalf of the Board of Directors"
    ws.cell(row=row, column=3).alignment = center_align
    row += 2
    ws.cell(row=row, column=1).value = "For M/s Siva Parvathi & Associates"
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row+1, column=1).value = "ICAI Firm registration number: 020872S"
    ws.cell(row=row+1, column=1).alignment = left_align
    ws.cell(row=row+2, column=1).value = "Chartered Accountants"
    ws.cell(row=row+2, column=1).alignment = left_align
    row += 4
    ws.cell(row=row, column=1).value = "S. Siva Parvathi"
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=3).value = "Director"
    ws.cell(row=row, column=3).alignment = center_align
    ws.cell(row=row, column=4).value = "Director"
    ws.cell(row=row, column=4).alignment = center_align
    row += 1
    ws.cell(row=row, column=1).value = "Proprietor"
    ws.cell(row=row, column=1).alignment = left_align
    row += 1
    ws.cell(row=row, column=1).value = "UDIN:24226087BKEECZ1200"
    ws.cell(row=row, column=1).alignment = left_align
    row += 1
    ws.cell(row=row, column=1).value = "Place: Hyderabad"
    ws.cell(row=row, column=1).alignment = left_align
    row += 1
    ws.cell(row=row, column=1).value = "Date: 04/09/2024"
    ws.cell(row=row, column=1).alignment = left_align

    # Save Excel file
    try:
        output_folder = "cashflow_excel"
        os.makedirs(output_folder, exist_ok=True)
        output_file = os.path.join(output_folder, "cashflow_report.xlsx")
        wb.save(output_file)
        print(f"Cash Flow Statement report generated successfully and saved to {output_file}")
        
    except PermissionError:
        print(f"PermissionError: Unable to save to {output_file}. Trying alternative location...")
        fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "cash_flow_report_fallback.xlsx")
        try:
            wb.save(fallback_file)
            print(f"Cash Flow Statement report saved to alternative location: {fallback_file}")
        except Exception as e:
            print(f"Failed to save Cash Flow Statement report: {str(e)}")
    except Exception as e:
        print(f"Error saving Cash Flow Statement report: {str(e)}")

if __name__ == "__main__":
    generate_cashflow_report()