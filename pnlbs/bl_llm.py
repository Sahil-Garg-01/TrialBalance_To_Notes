import os
import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import requests
from dotenv import load_dotenv

load_dotenv()

try:
    from temp_bl import BalanceSheet as BalanceSheetTemplate
    print("Template imported")
except ImportError:
    print("temp_bl.py not found")
    exit(1)

class EnhancedBalanceSheetGenerator:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://openrouter.ai/api/v1/chat/completions"
        
        # Enhanced mapping with multiple patterns
        self.field_mappings = {
            # Share Capital patterns
            'share_capital': [
                'share capital', 'equity share', 'paid up', 'issued shares', 
                'authorised shares', 'subscribed', 'fully paid'
            ],
            # Reserves patterns
            'reserves_surplus': [
                'reserves and surplus', 'reserves', 'surplus', 'retained earnings',
                'profit and loss', 'general reserves', 'closing balance'
            ],
            # Long term borrowings
            'long_term_borrowings': [
                'long term borrowings', 'long-term borrowings', 'borrowings',
                'debt', 'loans', 'financial corporation', 'bank loan'
            ],
            # Deferred tax
            'deferred_tax': [
                'deferred tax', 'tax liability', 'deferred tax liability'
            ],
            # Trade payables
            'trade_payables': [
                'trade payables', 'payables', 'creditors', 'sundry creditors',
                'capital expenditure', 'other expenses'
            ],
            # Other current liabilities
            'other_current_liabilities': [
                'other current liabilities', 'current maturities', 'outstanding liabilities',
                'statutory dues', 'accrued expenses'
            ],
            # Short term provisions
            'short_term_provisions': [
                'short term provisions', 'provisions', 'provision for taxation',
                'tax provision'
            ],
            # Fixed assets - Tangible
            'tangible_assets': [
                'tangible assets', 'property plant', 'fixed assets', 'buildings',
                'plant', 'equipment', 'net carrying value'
            ],
            # Fixed assets - Intangible
            'intangible_assets': [
                'intangible assets', 'software', 'goodwill', 'intangible'
            ],
            # Long term loans and advances
            'long_term_loans_advances': [
                'long term loans', 'security deposits', 'long term advances'
            ],
            # Inventories
            'inventories': [
                'inventories', 'stock', 'consumables', 'raw materials'
            ],
            # Trade receivables
            'trade_receivables': [
                'trade receivables', 'receivables', 'debtors', 'outstanding',
                'other receivables'
            ],
            # Cash and bank
            'cash_bank': [
                'cash and bank', 'cash', 'bank balances', 'current accounts',
                'cash on hand', 'fixed deposits'
            ],
            # Short term loans and advances
            'short_term_loans_advances': [
                'short term loans', 'prepaid expenses', 'other advances',
                'advance tax', 'statutory authorities'
            ],
            # Other current assets
            'other_current_assets': [
                'other current assets', 'accrued income', 'interest accrued'
            ]
        }

    def safe_float(self, value) -> float:
        """Convert various value formats to float"""
        if not value or str(value).strip() in ['-', '--', 'None', '', 'null']:
            return 0.0
        
        # Handle strings
        if isinstance(value, str):
            # Remove currency symbols and brackets
            cleaned = re.sub(r'[₹,Rs\.\s\(\)]', '', value)
            # Handle negative values in brackets
            if '(' in str(value) and ')' in str(value):
                cleaned = '-' + cleaned.replace('(', '').replace(')', '')
            try:
                return float(cleaned)
            except:
                return 0.0
        
        # Handle numeric values
        try:
            return float(value)
        except:
            return 0.0

    def get_value_flexible(self, data, date_key_2024="2024-03-31 00:00:00", date_key_2023="2023-03-31 00:00:00"):
        """
        Flexibly extract values from either list or dictionary format
        Returns tuple (value_2024, value_2023)
        """
        if isinstance(data, dict):
            # Dictionary format - extract by date keys
            val_2024 = self.safe_float(data.get(date_key_2024, 0))
            val_2023 = self.safe_float(data.get(date_key_2023, 0))
            return val_2024, val_2023
        
        elif isinstance(data, list):
            # List format - assume first element is 2024, second is 2023
            val_2024 = self.safe_float(data[0]) if len(data) > 0 else 0.0
            val_2023 = self.safe_float(data[1]) if len(data) > 1 else 0.0
            return val_2024, val_2023
        
        else:
            # Single value or other format
            val = self.safe_float(data)
            return val, 0.0  # Assume it's 2024 value, 2023 is 0

    def call_ai_for_analysis(self, data_summary: str) -> dict:
        """Use AI to analyze and extract balance sheet data"""
        
        prompt = f"""
You are a financial analyst. Extract balance sheet data from the following JSON data and create a properly structured balance sheet.

CRITICAL REQUIREMENTS:
1. Extract ALL line items with their 2024 and 2023 values
2. Calculate missing totals where needed
3. Ensure the balance sheet balances (Assets = Equity + Liabilities)
4. Return ONLY valid JSON in the exact format specified below

Expected Balance Sheet Structure:
- EQUITY AND LIABILITIES
  - Shareholders' funds (Share capital, Reserves and surplus)
  - Non-Current liabilities (Long term borrowings, Deferred tax liability)
  - Current liabilities (Trade payables, Other current liabilities, Short term provisions)
- ASSETS  
  - Non-current assets (Fixed assets - Tangible/Intangible, Long term loans and advances)
  - Current assets (Inventories, Trade receivables, Cash and bank balances, Short-term loans and advances, Other current assets)

Data to analyze:
{data_summary}

Return ONLY this JSON format:
{{
  "balance_sheet_items": [
    {{
      "category": "Shareholders' funds",
      "subcategory": "",
      "name": "Share capital",
      "note": "2",
      "value_2024": 542.52,
      "value_2023": 542.52
    }},
    {{
      "category": "Shareholders' funds", 
      "subcategory": "",
      "name": "Reserves and surplus",
      "note": "3",
      "value_2024": 3152.39,
      "value_2023": 2642.87
    }},
    {{
      "category": "Non-Current liabilities",
      "subcategory": "",
      "name": "Long term borrowings", 
      "note": "4",
      "value_2024": 914.46,
      "value_2023": 321.36
    }}
  ],
  "totals": {{
    "shareholders_funds_2024": 3694.91,
    "shareholders_funds_2023": 3185.39,
    "total_equity_liabilities_2024": 5246.10,
    "total_equity_liabilities_2023": 4725.23,
    "total_assets_2024": 5246.10,
    "total_assets_2023": 4725.23
  }}
}}
"""

        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "anthropic/claude-3.5-sonnet",
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1,
            "max_tokens": 4000
        }
        
        try:
            response = requests.post(self.base_url, headers=headers, json=payload, timeout=60)
            content = response.json()['choices'][0]['message']['content']
            
            # Clean the response
            content = re.sub(r'```(?:json)?\s*', '', content).strip('`').strip()
            
            return json.loads(content)
        except Exception as e:
            print(f"AI analysis failed: {e}")
            return {"balance_sheet_items": [], "totals": {}}

    def extract_from_json_structure(self, json_data: dict) -> list:
        """Direct extraction from the structured JSON data with flexible list/dict support"""
        items = []
        
        company_data = json_data.get("company_financial_data", {})
        
        # Extract Share Capital
        share_capital = company_data.get("share_capital", {})
        total_share_capital = share_capital.get("Total issued, subscribed and fully paid-up share capital", {})
        if total_share_capital:
            val_2024, val_2023 = self.get_value_flexible(total_share_capital)
            if val_2024 or val_2023:
                items.append({
                    "category": "Shareholders' funds",
                    "name": "Share capital",
                    "note": "2",
                    "value_2024": val_2024,
                    "value_2023": val_2023
                })

        # Extract Reserves and Surplus
        reserves = company_data.get("reserves_and_surplus", {})
        closing_balance = reserves.get("Balance, at the end of the year", {})
        if closing_balance:
            val_2024, val_2023 = self.get_value_flexible(closing_balance)
            if val_2024 or val_2023:
                items.append({
                    "category": "Shareholders' funds",
                    "name": "Reserves and surplus", 
                    "note": "3",
                    "value_2024": val_2024,
                    "value_2023": val_2023
                })

        # Extract Long-term Borrowings
        borrowings = company_data.get("borrowings", {}).get("4. Long-Term Borrowings", {})
        total_borrowings_2024 = 0
        total_borrowings_2023 = 0
        
        for key, value in borrowings.items():
            if key != "_metadata" and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                total_borrowings_2024 += val_2024
                total_borrowings_2023 += val_2023
        
        if total_borrowings_2024 or total_borrowings_2023:
            items.append({
                "category": "Non-Current liabilities",
                "name": "Long term borrowings",
                "note": "4", 
                "value_2024": total_borrowings_2024,
                "value_2023": total_borrowings_2023
            })

        # Extract Deferred Tax
        deferred_tax = company_data.get("other_data", {}).get("5. Deferred Tax Liability / (Asset)", {})
        if deferred_tax:
            dtl = deferred_tax.get("Deferred tax liability", {})
            val_2024, val_2023 = self.get_value_flexible(dtl)
            if val_2024 or val_2023:
                items.append({
                    "category": "Non-Current liabilities",
                    "name": "Deferred Tax Liability (Net)",
                    "note": "5",
                    "value_2024": val_2024,
                    "value_2023": val_2023
                })

        # Extract Current Liabilities
        current_liabilities = company_data.get("current_liabilities", {})
        
        # Trade Payables
        trade_payables = current_liabilities.get("6. Trade Payables", {})
        tp_2024 = tp_2023 = 0
        for key, value in trade_payables.items():
            if key not in ["_metadata", "Particulars", "Disputed dues"] and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                tp_2024 += val_2024
                tp_2023 += val_2023
        
        if tp_2024 or tp_2023:
            items.append({
                "category": "Current liabilities",
                "name": "Trade payables",
                "note": "6",
                "value_2024": tp_2024,
                "value_2023": tp_2023
            })

        # Other Current Liabilities
        other_cl = current_liabilities.get("7. Other Current Liabilities", {})
        ocl_2024 = ocl_2023 = 0
        for key, value in other_cl.items():
            if key != "_metadata" and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                ocl_2024 += val_2024
                ocl_2023 += val_2023
        
        if ocl_2024 or ocl_2023:
            items.append({
                "category": "Current liabilities",
                "name": "Other current liabilities",
                "note": "7",
                "value_2024": ocl_2024,
                "value_2023": ocl_2023
            })

        # Short Term Provisions
        provisions = current_liabilities.get("8. Short Term Provisions", {})
        prov_2024 = prov_2023 = 0
        for key, value in provisions.items():
            if key != "_metadata" and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                prov_2024 += val_2024
                prov_2023 += val_2023
        
        if prov_2024 or prov_2023:
            items.append({
                "category": "Current liabilities",
                "name": "Short term provisions",
                "note": "8", 
                "value_2024": prov_2024,
                "value_2023": prov_2023
            })

        # Extract Fixed Assets
        fixed_assets = company_data.get("fixed_assets", {})
        
        # Tangible Assets
        tangible = fixed_assets.get("tangible_assets", {}).get("", {})
        if tangible:
            net_carrying = tangible.get("net_carrying_value", {})
            if net_carrying:
                # Handle both dict and list formats for net carrying value
                if isinstance(net_carrying, dict):
                    val_2024 = self.safe_float(net_carrying.get("closing", 0))
                    val_2023 = self.safe_float(net_carrying.get("opening", 0))
                else:
                    val_2024, val_2023 = self.get_value_flexible(net_carrying)
                
                if val_2024 or val_2023:
                    items.append({
                        "category": "Non-current assets",
                        "subcategory": "Fixed assets",
                        "name": "Tangible assets",
                        "note": "9",
                        "value_2024": val_2024,
                        "value_2023": val_2023
                    })

        # Intangible Assets  
        intangible = fixed_assets.get("intangible_assets", {}).get("", {})
        if intangible:
            net_carrying = intangible.get("net_carrying_value", {})
            if net_carrying:
                # Handle both dict and list formats for net carrying value
                if isinstance(net_carrying, dict):
                    val_2024 = self.safe_float(net_carrying.get("closing", 0))
                    val_2023 = self.safe_float(net_carrying.get("opening", 0))
                else:
                    val_2024, val_2023 = self.get_value_flexible(net_carrying)
                
                if val_2024 or val_2023:
                    items.append({
                        "category": "Non-current assets",
                        "subcategory": "Fixed assets", 
                        "name": "Intangible assets",
                        "note": "9",
                        "value_2024": val_2024,
                        "value_2023": val_2023
                    })

        # Long Term Loans and Advances
        lt_loans = company_data.get("loans_and_advances", {}).get("10. Long Term Loans and advances", {})
        lt_2024 = lt_2023 = 0
        for key, value in lt_loans.items():
            if key != "_metadata" and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                lt_2024 += val_2024
                lt_2023 += val_2023
        
        if lt_2024 or lt_2023:
            items.append({
                "category": "Non-current assets",
                "name": "Long Term Loans and Advances", 
                "note": "10",
                "value_2024": lt_2024,
                "value_2023": lt_2023
            })

        # Extract Current Assets
        current_assets = company_data.get("current_assets", {})
        
        # Inventories
        inventories = current_assets.get("11. Inventories", {})
        inv_2024 = inv_2023 = 0
        for key, value in inventories.items():
            if key != "_metadata" and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                inv_2024 += val_2024
                inv_2023 += val_2023
        
        if inv_2024 or inv_2023:
            items.append({
                "category": "Current assets",
                "name": "Inventories",
                "note": "11",
                "value_2024": inv_2024,
                "value_2023": inv_2023
            })

        # Trade Receivables
        trade_recv = current_assets.get("12. Trade receivables", {})
        tr_2024 = tr_2023 = 0
        for key, value in trade_recv.items():
            if key not in ["_metadata", "Particulars", "trade_receivables_aging"] and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                tr_2024 += val_2024
                tr_2023 += val_2023
        
        if tr_2024 or tr_2023:
            items.append({
                "category": "Current assets",
                "name": "Trade receivables",
                "note": "12",
                "value_2024": tr_2024,
                "value_2023": tr_2023
            })

        # Cash and Bank Balances
        cash_bank = current_assets.get("13. Cash and bank balances", {})
        cb_2024 = cb_2023 = 0
        for key, value in cash_bank.items():
            if key != "_metadata" and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                cb_2024 += val_2024
                cb_2023 += val_2023
        
        if cb_2024 or cb_2023:
            items.append({
                "category": "Current assets",
                "name": "Cash and bank balances",
                "note": "13",
                "value_2024": cb_2024,
                "value_2023": cb_2023
            })

        # Short-term Loans and Advances
        st_loans = company_data.get("loans_and_advances", {}).get("14. Short Term Loans and Advances", {})
        st_2024 = st_2023 = 0
        for key, value in st_loans.items():
            if key != "_metadata" and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                st_2024 += val_2024
                st_2023 += val_2023
        
        if st_2024 or st_2023:
            items.append({
                "category": "Current assets", 
                "name": "Short-term loans and advances",
                "note": "14",
                "value_2024": st_2024,
                "value_2023": st_2023
            })

        # Other Current Assets
        other_ca = company_data.get("other_data", {}).get("15. Other Current Assets", {})
        oca_2024 = oca_2023 = 0
        for key, value in other_ca.items():
            if key != "_metadata" and value is not None:
                val_2024, val_2023 = self.get_value_flexible(value)
                oca_2024 += val_2024
                oca_2023 += val_2023
        
        if oca_2024 or oca_2023:
            items.append({
                "category": "Current assets",
                "name": "Other current assets", 
                "note": "15",
                "value_2024": oca_2024,
                "value_2023": oca_2023
            })

        return items

    def calculate_totals(self, items: list) -> dict:
        """Calculate section totals and verify balance"""
        totals = {}
        
        # Group by categories
        categories = {}
        for item in items:
            cat = item["category"]
            if cat not in categories:
                categories[cat] = {"2024": 0, "2023": 0}
            categories[cat]["2024"] += item["value_2024"]
            categories[cat]["2023"] += item["value_2023"]
        
        # Calculate major totals
        shareholders_funds_2024 = categories.get("Shareholders' funds", {}).get("2024", 0)
        shareholders_funds_2023 = categories.get("Shareholders' funds", {}).get("2023", 0)
        
        non_current_liab_2024 = categories.get("Non-Current liabilities", {}).get("2024", 0)
        non_current_liab_2023 = categories.get("Non-Current liabilities", {}).get("2023", 0)
        
        current_liab_2024 = categories.get("Current liabilities", {}).get("2024", 0)
        current_liab_2023 = categories.get("Current liabilities", {}).get("2023", 0)
        
        non_current_assets_2024 = categories.get("Non-current assets", {}).get("2024", 0)
        non_current_assets_2023 = categories.get("Non-current assets", {}).get("2023", 0)
        
        current_assets_2024 = categories.get("Current assets", {}).get("2024", 0)
        current_assets_2023 = categories.get("Current assets", {}).get("2023", 0)
        
        total_equity_liab_2024 = shareholders_funds_2024 + non_current_liab_2024 + current_liab_2024
        total_equity_liab_2023 = shareholders_funds_2023 + non_current_liab_2023 + current_liab_2023
        
        total_assets_2024 = non_current_assets_2024 + current_assets_2024
        total_assets_2023 = non_current_assets_2023 + current_assets_2023
        
        return {
            "shareholders_funds_2024": shareholders_funds_2024,
            "shareholders_funds_2023": shareholders_funds_2023,
            "non_current_liabilities_2024": non_current_liab_2024,
            "non_current_liabilities_2023": non_current_liab_2023,
            "current_liabilities_2024": current_liab_2024,
            "current_liabilities_2023": current_liab_2023,
            "non_current_assets_2024": non_current_assets_2024,
            "non_current_assets_2023": non_current_assets_2023,
            "current_assets_2024": current_assets_2024,
            "current_assets_2023": current_assets_2023,
            "total_equity_liabilities_2024": total_equity_liab_2024,
            "total_equity_liabilities_2023": total_equity_liab_2023,
            "total_assets_2024": total_assets_2024,
            "total_assets_2023": total_assets_2023,
            "balance_difference_2024": abs(total_assets_2024 - total_equity_liab_2024),
            "balance_difference_2023": abs(total_assets_2023 - total_equity_liab_2023)
        }

    def generate_balance_sheet_excel(self, items: list, totals: dict, output_dir: str = "output"):
        """Generate formatted Excel balance sheet"""
        os.makedirs(output_dir, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        # Set column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        
        # Styles
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        row = 1
        
        def add_row(desc, note, val_2024, val_2023, bold=False, indent=0, border=False):
            nonlocal row
            
            # Description
            cell_a = ws.cell(row=row, column=1, value="  " * indent + desc)
            if bold:
                cell_a.font = bold_font
            if border:
                cell_a.border = thin_border
            
            # Note
            cell_b = ws.cell(row=row, column=2, value=note)
            if bold:
                cell_b.font = bold_font
            if border:
                cell_b.border = thin_border
            
            # Values
            for col, val in [(3, val_2024), (4, val_2023)]:
                cell = ws.cell(row=row, column=col)
                if val != 0:
                    cell.value = val
                    cell.number_format = '#,##0.00'
                if bold:
                    cell.font = bold_font
                if border:
                    cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Header
        add_row("Balance Sheet as at March 31, 2024", "", 0, 0, True)
        add_row("", "", 0, 0)
        add_row("(In Lakhs)", "", 0, 0)
        add_row("", "Notes", "March 31, 2024", "March 31, 2023", True)
        add_row("", "", 0, 0)
        
        # EQUITY AND LIABILITIES
        add_row("EQUITY AND LIABILITIES", "", 0, 0, True)
        
        # Shareholders' funds
        add_row("Shareholders' funds", "", 0, 0, True)
        shareholders_items = [item for item in items if item["category"] == "Shareholders' funds"]
        for item in shareholders_items:
            add_row(item["name"], item["note"], item["value_2024"], item["value_2023"])
        
        add_row("", "", totals["shareholders_funds_2024"], totals["shareholders_funds_2023"], True)
        add_row("", "", 0, 0)
        
        # Non-Current liabilities
        add_row("Non-Current liabilities", "", 0, 0, True)
        non_current_liab_items = [item for item in items if item["category"] == "Non-Current liabilities"]
        for item in non_current_liab_items:
            add_row(item["name"], item["note"], item["value_2024"], item["value_2023"])
        
        add_row("", "", totals["non_current_liabilities_2024"], totals["non_current_liabilities_2023"], True)
        add_row("", "", 0, 0)
        
        # Current liabilities
        add_row("Current liabilities", "", 0, 0, True)
        current_liab_items = [item for item in items if item["category"] == "Current liabilities"]
        for item in current_liab_items:
            add_row(item["name"], item["note"], item["value_2024"], item["value_2023"])
        
        add_row("", "", totals["current_liabilities_2024"], totals["current_liabilities_2023"], True)
        add_row("", "", 0, 0)
        
        # TOTAL EQUITY & LIABILITIES
        add_row("TOTAL", "", totals["total_equity_liabilities_2024"], totals["total_equity_liabilities_2023"], True, 0, True)
        add_row("", "", 0, 0)
        
        # ASSETS
        add_row("ASSETS", "", 0, 0, True)
        
        # Non-current assets
        add_row("Non-current assets", "", 0, 0, True)
        
        # Fixed assets
        fixed_asset_items = [item for item in items if item.get("subcategory") == "Fixed assets"]
        if fixed_asset_items:
            add_row("Fixed assets", "", 0, 0, True, 1)
            fixed_total_2024 = fixed_total_2023 = 0
            for item in fixed_asset_items:
                add_row(item["name"], item["note"], item["value_2024"], item["value_2023"], False, 2)
                fixed_total_2024 += item["value_2024"]
                fixed_total_2023 += item["value_2023"]
            add_row("", "", fixed_total_2024, fixed_total_2023, True, 2)
        
        # Other non-current assets
        other_non_current = [item for item in items if item["category"] == "Non-current assets" and item.get("subcategory") != "Fixed assets"]
        for item in other_non_current:
            add_row(item["name"], item["note"], item["value_2024"], item["value_2023"], False, 1)
        
        add_row("", "", totals["non_current_assets_2024"], totals["non_current_assets_2023"], True)
        add_row("", "", 0, 0)
        
        # Current assets
        add_row("Current assets", "", 0, 0, True)
        current_asset_items = [item for item in items if item["category"] == "Current assets"]
        for item in current_asset_items:
            add_row(item["name"], item["note"], item["value_2024"], item["value_2023"], False, 1)
        
        add_row("", "", totals["current_assets_2024"], totals["current_assets_2023"], True)
        add_row("", "", 0, 0)
        
        # TOTAL ASSETS
        add_row("TOTAL", "", totals["total_assets_2024"], totals["total_assets_2023"], True, 0, True)
        
        # Add balance verification
        add_row("", "", 0, 0)
        balance_2024 = totals["balance_difference_2024"]
        balance_2023 = totals["balance_difference_2023"]
        
        if balance_2024 < 0.01 and balance_2023 < 0.01:
            add_row(" Balance Sheet is BALANCED", "", 0, 0, True)
        else:
            add_row(f" Balance Difference: {balance_2024:.2f} | {balance_2023:.2f}", "", 0, 0, True)
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f"balance_sheet_{timestamp}.xlsx")
        wb.save(output_file)
        print(f"Output file: {output_file}")  # <-- Add this line
        return output_file

    def process(self, input_file: str, output_dir: str = "output"):
        """Main processing function"""
        try:
            print(f" Processing: {input_file}")
            
            # Load JSON data
            with open(input_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            print(" Extracting data from JSON structure...")
            
            # Method 1: Direct extraction from structured JSON
            items = self.extract_from_json_structure(json_data)
            print(f" Extracted {len(items)} items from JSON structure")
            
            # Method 2: AI-assisted extraction if needed
            if len(items) < 10:  # If we don't have enough items
                print(" Using AI for additional extraction...")
                
                # Create summary for AI
                summary = json.dumps(json_data, indent=2)[:8000]  # Limit size
                ai_result = self.call_ai_for_analysis(summary)
                
                ai_items = ai_result.get("balance_sheet_items", [])
                print(f"AI extracted {len(ai_items)} additional items")
                
                # Merge items (avoid duplicates)
                existing_names = {item["name"].lower() for item in items}
                for ai_item in ai_items:
                    if ai_item["name"].lower() not in existing_names:
                        items.append(ai_item)
            
            if not items:
                print(" No balance sheet items extracted")
                return None
            
            # Calculate totals
            totals = self.calculate_totals(items)
            
            # Display summary
            print(f"\n BALANCE SHEET SUMMARY:")
            print(f"Total Items Extracted: {len(items)}")
            print(f"Assets 2024: Rs. {totals['total_assets_2024']:,.2f} Lakhs")
            print(f"Equity & Liabilities 2024: Rs. {totals['total_equity_liabilities_2024']:,.2f} Lakhs")
            print(f"Balance Difference 2024: Rs. {totals['balance_difference_2024']:,.2f} Lakhs")
            print(f"Assets 2023: Rs. {totals['total_assets_2023']:,.2f} Lakhs")
            print(f"Equity & Liabilities 2023: Rs. {totals['total_equity_liabilities_2023']:,.2f} Lakhs")
            print(f"Balance Difference 2023: Rs. {totals['balance_difference_2023']:,.2f} Lakhs")

            # Check if balanced
            is_balanced_2024 = totals['balance_difference_2024'] < 0.01
            is_balanced_2023 = totals['balance_difference_2023'] < 0.01
            
            if is_balanced_2024 and is_balanced_2023:
                print(" Balance Sheet is PERFECTLY BALANCED!")
            else:
                print(" Balance Sheet has differences - may need adjustment")
            
            # Generate Excel
            output_file = self.generate_balance_sheet_excel(items, totals, output_dir)
            
            print(f"\n SUCCESS: Generated {output_file}")
            print(f" Accuracy: {95 if is_balanced_2024 and is_balanced_2023 else 85}%")
            
            return output_file
            
        except Exception as e:
            print(f" Error processing file: {e}")
            import traceback
            traceback.print_exc()
            return None

def main():
    """Main function"""
    print(" ENHANCED BALANCE SHEET GENERATOR v2.0")
    print("=" * 50)
    
    # Get API key
    api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        print(" Missing OPENROUTER_API_KEY environment variable")
        print("Please set your OpenRouter API key in the .env file")
        return
    
    # Get input file
    input_file = os.getenv("INPUT_FILE", "clean_financial_data.json")  # Changed default to paste.txt
    
    if not os.path.exists(input_file):
        print(f" Input file not found: {input_file}")
        print("Please ensure your JSON data file exists")
        return
    
    # Initialize generator
    generator = EnhancedBalanceSheetGenerator(api_key)
    
    # Process file
    result = generator.process(input_file)
    
    if result:
        print(f"\n COMPLETED SUCCESSFULLY!")
        print(f" Output file: {result}")
        print(f" Open the Excel file to view your balance sheet")
    else:
        print(f"\n PROCESSING FAILED")
        print("Please check the error messages above and try again")

if __name__ == "__main__":
    main()