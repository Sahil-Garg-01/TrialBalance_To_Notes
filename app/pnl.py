import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_note_data(note_number, folder="generated_notes"):
    """
    Load note data for a specific note_number from notes.json in the specified folder.
    Compatible with {"notes": [ ... ]} structure.
    """
    file_path = os.path.join(folder, "notes.json")
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            notes = data.get("notes", [])
            for note in notes:
                # Try both direct and metadata field for note_number
                n_num = note.get("note_number") or note.get("metadata", {}).get("note_number")
                if str(n_num) == str(note_number):
                    return note
            print(f"Warning: Note {note_number} not found in {file_path}")
            return None
    except FileNotFoundError:
        print(f"Warning: Note {note_number} file not found at {file_path}")
        return None
    except json.JSONDecodeError:
        print(f"Warning: Invalid JSON in note {note_number}")
        return None

def extract_total_from_note(note_data, year="2024"):
    """Extract total value from note data for the specified year."""
    if not note_data or "structure" not in note_data:
        return 0.0
    
    # Look for total in categories
    for category in note_data["structure"]:
        if year == "2024" and "total" in category:
            try:
                return float(category["total"])
            except (ValueError, TypeError):
                continue
        elif year == "2023" and "previous_total" in category:
            try:
                return float(category["previous_total"])
            except (ValueError, TypeError):
                continue
    
    # If no total found, sum up subcategory values
    total = 0.0
    for category in note_data["structure"]:
        if "subcategories" in category:
            for subcat in category["subcategories"]:
                if year == "2024" and "value" in subcat:
                    try:
                        total += float(subcat["value"])
                    except (ValueError, TypeError):
                        continue
                elif year == "2023" and "previous_value" in subcat:
                    try:
                        total += float(subcat["previous_value"])
                    except (ValueError, TypeError):
                        continue
    
    return total

def extract_specific_value(note_data, key, year="2024"):
    """Extract specific value from note data based on key and year."""
    if not note_data or "structure" not in note_data:
        return 0.0
    
    # Search through all categories and subcategories
    for category in note_data["structure"]:
        if "subcategories" in category:
            for subcat in category["subcategories"]:
                label = subcat.get("label", "").lower().replace(" ", "").replace("-", "").replace("/", "").replace("&", "")
                if key.lower() in label:
                    if year == "2024" and "value" in subcat:
                        try:
                            return float(subcat["value"])
                        except (ValueError, TypeError):
                            continue
                    elif year == "2023" and "previous_value" in subcat:
                        try:
                            return float(subcat["previous_value"])
                        except (ValueError, TypeError):
                            continue
    
    return 0.0

def format_currency(value):
    """Format currency value for display."""
    if isinstance(value, (int, float)) and value != 0:
        return f"{value:,.2f}"
    return "0.00"

def generate_pnl_report():
    """Generate P&L report in Excel format using data from generated_notes folder."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit and Loss Statement"

    # Define styles
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")

    # Set column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Header
    ws["A1"] = "Profit and Loss Statement for the Years Ended March 31, 2024 and 2023"
    ws["A1"].font = bold_font
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = center_align

    # Table headers
    headers = ["Particulars", "Note No.", "Year Ended 31.03.2024", "Year Ended 31.03.2023"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align if col > 1 else left_align

    # Load all required notes
    notes_data = {}
    for note_num in range(16, 29):  # Notes 16-28
        notes_data[str(note_num)] = load_note_data(str(note_num))

    # Calculate values
    revenue_2024 = extract_total_from_note(notes_data.get("16"), "2024")
    revenue_2023 = extract_total_from_note(notes_data.get("16"), "2023")
    other_income_2024 = extract_total_from_note(notes_data.get("17"), "2024")
    other_income_2023 = extract_total_from_note(notes_data.get("17"), "2023")
    
    total_income_2024 = revenue_2024 + other_income_2024
    total_income_2023 = revenue_2023 + other_income_2023

    # Calculate expenses
    cost_materials_2024 = extract_total_from_note(notes_data.get("18"), "2024")
    cost_materials_2023 = extract_total_from_note(notes_data.get("18"), "2023")
    employee_benefit_2024 = extract_total_from_note(notes_data.get("19"), "2024")
    employee_benefit_2023 = extract_total_from_note(notes_data.get("19"), "2023")
    other_expenses_2024 = extract_total_from_note(notes_data.get("20"), "2024")
    other_expenses_2023 = extract_total_from_note(notes_data.get("20"), "2023")
    depreciation_2024 = extract_total_from_note(notes_data.get("21"), "2024")
    depreciation_2023 = extract_total_from_note(notes_data.get("21"), "2023")
    loss_assets_2024 = extract_total_from_note(notes_data.get("22"), "2024")
    loss_assets_2023 = extract_total_from_note(notes_data.get("22"), "2023")
    finance_costs_2024 = extract_total_from_note(notes_data.get("23"), "2024")
    finance_costs_2023 = extract_total_from_note(notes_data.get("23"), "2023")
    auditor_payment_2024 = extract_total_from_note(notes_data.get("24"), "2024")
    auditor_payment_2023 = extract_total_from_note(notes_data.get("24"), "2023")

    total_expenses_2024 = (cost_materials_2024 + employee_benefit_2024 + other_expenses_2024 + 
                          depreciation_2024 + loss_assets_2024 + finance_costs_2024 + auditor_payment_2024)
    total_expenses_2023 = (cost_materials_2023 + employee_benefit_2023 + other_expenses_2023 + 
                          depreciation_2023 + loss_assets_2023 + finance_costs_2023 + auditor_payment_2023)

    profit_before_exceptional_2024 = total_income_2024 - total_expenses_2024
    profit_before_exceptional_2023 = total_income_2023 - total_expenses_2023

    # Tax calculations (assuming notes 25-28 contain tax information)
    current_tax_2024 = extract_total_from_note(notes_data.get("25"), "2024")
    current_tax_2023 = extract_total_from_note(notes_data.get("25"), "2023")
    deferred_tax_2024 = extract_total_from_note(notes_data.get("26"), "2024")
    deferred_tax_2023 = extract_total_from_note(notes_data.get("26"), "2023")

    total_tax_2024 = current_tax_2024 + deferred_tax_2024
    total_tax_2023 = current_tax_2023 + deferred_tax_2023

    profit_after_tax_2024 = profit_before_exceptional_2024 - total_tax_2024
    profit_after_tax_2023 = profit_before_exceptional_2023 - total_tax_2023

    # P&L line items with calculated values
    line_items = [
        {"label": "I. Revenue from Operations", "note": "16", "value_2024": revenue_2024, "value_2023": revenue_2023},
        {"label": "II. Other Income", "note": "17", "value_2024": other_income_2024, "value_2023": other_income_2023},
        {"label": "III. Total Income (I + II)", "note": "", "value_2024": total_income_2024, "value_2023": total_income_2023},
        {"label": "IV. Expenses", "note": "", "value_2024": "", "value_2023": ""},
        {"label": "   Cost of Materials Consumed", "note": "18", "value_2024": cost_materials_2024, "value_2023": cost_materials_2023},
        {"label": "   Employee Benefit Expense", "note": "19", "value_2024": employee_benefit_2024, "value_2023": employee_benefit_2023},
        {"label": "   Other Expenses", "note": "20", "value_2024": other_expenses_2024, "value_2023": other_expenses_2023},
        {"label": "   Depreciation and Amortisation Expense", "note": "21", "value_2024": depreciation_2024, "value_2023": depreciation_2023},
        {"label": "   Loss on Sale of Assets & Investments", "note": "22", "value_2024": loss_assets_2024, "value_2023": loss_assets_2023},
        {"label": "   Finance Costs", "note": "23", "value_2024": finance_costs_2024, "value_2023": finance_costs_2023},
        {"label": "   Payment to Auditor", "note": "24", "value_2024": auditor_payment_2024, "value_2023": auditor_payment_2023},
        {"label": "   Total Expenses", "note": "", "value_2024": total_expenses_2024, "value_2023": total_expenses_2023},
        {"label": "V. Profit Before Exceptional and Extraordinary Items and Tax (III - IV)", "note": "", "value_2024": profit_before_exceptional_2024, "value_2023": profit_before_exceptional_2023},
        {"label": "VI. Exceptional Items", "note": "27", "value_2024": extract_total_from_note(notes_data.get("27"), "2024"), "value_2023": extract_total_from_note(notes_data.get("27"), "2023")},
        {"label": "VII. Profit Before Tax (V - VI)", "note": "", "value_2024": profit_before_exceptional_2024, "value_2023": profit_before_exceptional_2023},
        {"label": "VIII. Tax Expense:", "note": "", "value_2024": "", "value_2023": ""},
        {"label": "   (1) Current Tax", "note": "25", "value_2024": current_tax_2024, "value_2023": current_tax_2023},
        {"label": "   (2) Deferred Tax", "note": "26", "value_2024": deferred_tax_2024, "value_2023": deferred_tax_2023},
        {"label": "IX. Profit After Tax for the period (VII - VIII)", "note": "", "value_2024": profit_after_tax_2024, "value_2023": profit_after_tax_2023}
    ]

    # Write line items to Excel
    row = 4
    for item in line_items:
        ws.cell(row=row, column=1).value = item["label"]
        ws.cell(row=row, column=2).value = item["note"]
        
        # Format values
        if item["value_2024"] == "":
            ws.cell(row=row, column=3).value = ""
        else:
            ws.cell(row=row, column=3).value = format_currency(item["value_2024"])
            
        if item["value_2023"] == "":
            ws.cell(row=row, column=4).value = ""
        else:
            ws.cell(row=row, column=4).value = format_currency(item["value_2023"])
        
        # Apply formatting
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align if col > 1 else left_align
            if item["label"].startswith(("I.", "II.", "III.", "IV.", "V.", "VI.", "VII.", "VIII.", "IX.")):
                ws.cell(row=row, column=col).font = bold_font
        row += 1

    # Add notes and disclosures
    row += 1
    ws.cell(row=row, column=1).value = "Notes:"
    ws.cell(row=row, column=1).font = bold_font
    row += 1

    # Add any specific disclosures from notes
    note_20_data = notes_data.get("20")
    if note_20_data and "notes_and_disclosures" in note_20_data:
        ws.cell(row=row, column=1).value = note_20_data["notes_and_disclosures"][0]
        ws.cell(row=row, column=1).alignment = left_align
        row += 1

    # EPS information (if available in notes 27-28)
    eps_data = notes_data.get("28")  # Assuming EPS is in note 28
    if eps_data:
        ws.cell(row=row, column=1).value = "Earnings Per Share (EPS):"
        ws.cell(row=row, column=1).font = bold_font
        row += 1
        
        basic_eps_2024 = extract_specific_value(eps_data, "basic", "2024")
        basic_eps_2023 = extract_specific_value(eps_data, "basic", "2023")
        diluted_eps_2024 = extract_specific_value(eps_data, "diluted", "2024")
        diluted_eps_2023 = extract_specific_value(eps_data, "diluted", "2023")
        
        ws.cell(row=row, column=1).value = "   Basic EPS"
        ws.cell(row=row, column=3).value = format_currency(basic_eps_2024)
        ws.cell(row=row, column=4).value = format_currency(basic_eps_2023)
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=4).alignment = center_align
        row += 1
        
        ws.cell(row=row, column=1).value = "   Diluted EPS"
        ws.cell(row=row, column=3).value = format_currency(diluted_eps_2024)
        ws.cell(row=row, column=4).value = format_currency(diluted_eps_2023)
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=4).alignment = center_align

    # Save Excel file with error handling
    output_folder = "pnl_excel"
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, "pnl_report.xlsx")
    try:
        wb.save(output_file)
        print(f"P&L report generated successfully and saved to {output_file}")
        
    except PermissionError:
        print(f"PermissionError: Unable to save to {output_file}. Trying alternative location...")
        fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "pnl_report_fallback.xlsx")
        try:
            wb.save(fallback_file)
            print(f"P&L report saved to alternative location: {fallback_file}")
        except Exception as e:
            print(f"Failed to save P&L report: {str(e)}")
    except Exception as e:
        print(f"Error saving P&L report: {str(e)}")

if __name__ == "__main__":
    generate_pnl_report()