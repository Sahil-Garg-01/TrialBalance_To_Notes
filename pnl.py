import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_note_data(note_number, folder="generated_notes"):
    """Load note data from JSON file in the specified folder."""
    file_path = os.path.join(folder, f"note_{note_number}.json")
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return None

def extract_value(note_data, key, default=None):
    """Extract a value from note data, handling placeholders or actual values for 2024 and 2023."""
    if not note_data:
        return default or f"{{{key}}}"
    # Check subcategories first
    for category in note_data.get("structure", []):
        for subcat in category.get("subcategories", []):
            if subcat.get("value") and key.startswith(subcat.get("label", "").lower().replace(" ", "").replace("-", "").replace("/", "").replace("&", "")) and key.endswith("_2024"):
                return subcat.get("value", default or f"{{{key}}}")
            if subcat.get("previous_value") and key.startswith(subcat.get("label", "").lower().replace(" ", "").replace("-", "").replace("/", "").replace("&", "")) and key.endswith("_2023"):
                return subcat.get("previous_value", default or f"{{{key}}}")
    # Check for total and previous_total across all categories
    for category in note_data.get("structure", []):
        if "total" in category and key.endswith("_total_2024"):
            return category.get("total", default or f"{{{key}}}")
        if "previous_total" in category and key.endswith("_total_2023"):
            return category.get("previous_total", default or f"{{{key}}}")
    return default or f"{{{key}}}"

def generate_pnl_report():
    """Generate P&L report in Excel format using data from generated_notes folder."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit and Loss Statement"

    # Define styles
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")

    # Set column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Header
    ws["A1"] = "Profit and Loss Statement for the Years Ended March 31, 2024 and 2023"
    ws["A1"].font = bold_font
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = center_align

    # Table headers
    headers = ["Particulars", "Note No.", "Year Ended 31.03.2024", "Year Ended 31.03.2023"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align if col > 1 else left_align

    # P&L line items
    row = 4
    line_items = [
        {"label": "I. Revenue from Operations", "note": "16", "key_2024": "total_2024", "key_2023": "total_2023", "value_2024": None, "value_2023": None},
        {"label": "II. Other Income", "note": "17", "key_2024": "total_2024", "key_2023": "total_2023", "value_2024": None, "value_2023": None},
        {"label": "III. Total Income (I + II)", "note": "", "key_2024": "total_income_2024", "key_2023": "total_income_2023", "value_2024": None, "value_2023": None},
        {"label": "IV. Expenses", "note": "", "key_2024": "", "key_2023": "", "value_2024": None, "value_2023": None},
        {"label": "   Cost of Materials Consumed", "note": "18", "key_2024": "costmaterialsconsumed_2024", "key_2023": "costmaterialsconsumed_2023", "value_2024": None, "value_2023": None},
        {"label": "   Employee Benefit Expense", "note": "19", "key_2024": "total_2024", "key_2023": "total_2023", "value_2024": None, "value_2023": None},
        {"label": "   Other Expenses", "note": "20", "key_2024": "total_2024", "key_2023": "total_2023", "value_2024": None, "value_2023": None},
        {"label": "   Depreciation and Amortisation Expense", "note": "21", "key_2024": "total_2024", "key_2023": "total_2023", "value_2024": None, "value_2023": None},
        {"label": "   Loss on Sale of Assets & Investments", "note": "22", "key_2024": "total_2024", "key_2023": "total_2023", "value_2024": None, "value_2023": None},
        {"label": "   Finance Costs", "note": "23", "key_2024": "total_2024", "key_2023": "total_2023", "value_2024": None, "value_2023": None},
        {"label": "   Payment to Auditor", "note": "24", "key_2024": "total_2024", "key_2023": "total_2023", "value_2024": None, "value_2023": None},
        {"label": "   Total Expenses", "note": "", "key_2024": "total_expenses_2024", "key_2023": "total_expenses_2023", "value_2024": None, "value_2023": None},
        {"label": "V. Profit Before Exceptional and Extraordinary Items and Tax (III - IV)", "note": "", "key_2024": "profit_before_exceptional_2024", "key_2023": "profit_before_exceptional_2023", "value_2024": None, "value_2023": None},
        {"label": "VI. Exceptional Items", "note": "", "key_2024": "exceptional_items_2024", "key_2023": "exceptional_items_2023", "value_2024": "{exceptional_items_2024}", "value_2023": "{exceptional_items_2023}"},
        {"label": "VII. Profit Before Tax (V - VI)", "note": "", "key_2024": "profit_before_tax_2024", "key_2023": "profit_before_tax_2023", "value_2024": None, "value_2023": None},
        {"label": "VIII. Tax Expense:", "note": "", "key_2024": "", "key_2023": "", "value_2024": None, "value_2023": None},
        {"label": "   (1) Current Tax", "note": "", "key_2024": "current_tax_2024", "key_2023": "current_tax_2023", "value_2024": "{current_tax_2024}", "value_2023": "{current_tax_2023}"},
        {"label": "   (2) Deferred Tax", "note": "", "key_2024": "deferred_tax_2024", "key_2023": "deferred_tax_2023", "value_2024": "{deferred_tax_2024}", "value_2023": "{deferred_tax_2023}"},
        {"label": "IX. Profit After Tax for the period (VII - VIII)", "note": "", "key_2024": "profit_after_tax_2024", "key_2023": "profit_after_tax_2023", "value_2024": None, "value_2023": None}
    ]

    # Load note data and populate values
    note_values = {}
    for note_number in [str(i) for i in range(16, 25)]:
        note_data = load_note_data(note_number)
        if note_data:
            for item in line_items:
                if item["note"] == note_number:
                    item["value_2024"] = extract_value(note_data, item["key_2024"])
                    item["value_2023"] = extract_value(note_data, item["key_2023"])

    # Calculate totals for 2024 and 2023
    revenue_2024 = line_items[0]["value_2024"]
    revenue_2023 = line_items[0]["value_2023"]
    other_income_2024 = line_items[1]["value_2024"]
    other_income_2023 = line_items[1]["value_2023"]
    total_income_2024 = "{total_income_2024}" if isinstance(revenue_2024, str) or isinstance(other_income_2024, str) else revenue_2024 + other_income_2024 if revenue_2024 and other_income_2024 else "{total_income_2024}"
    total_income_2023 = "{total_income_2023}" if isinstance(revenue_2023, str) or isinstance(other_income_2023, str) else revenue_2023 + other_income_2023 if revenue_2023 and other_income_2023 else "{total_income_2023}"
    line_items[2]["value_2024"] = total_income_2024
    line_items[2]["value_2023"] = total_income_2023

    expenses_2024 = [item["value_2024"] for item in line_items[4:11]]
    expenses_2023 = [item["value_2023"] for item in line_items[4:11]]
    total_expenses_2024 = "{total_expenses_2024}"
    total_expenses_2023 = "{total_expenses_2023}"
    if all(isinstance(x, (int, float)) for x in expenses_2024):
        total_expenses_2024 = sum(expenses_2024)
    if all(isinstance(x, (int, float)) for x in expenses_2023):
        total_expenses_2023 = sum(expenses_2023)
    line_items[11]["value_2024"] = total_expenses_2024
    line_items[11]["value_2023"] = total_expenses_2023

    profit_before_exceptional_2024 = "{profit_before_exceptional_2024}"
    profit_before_exceptional_2023 = "{profit_before_exceptional_2023}"
    if isinstance(total_income_2024, (int, float)) and isinstance(total_expenses_2024, (int, float)):
        profit_before_exceptional_2024 = total_income_2024 - total_expenses_2024
    if isinstance(total_income_2023, (int, float)) and isinstance(total_expenses_2023, (int, float)):
        profit_before_exceptional_2023 = total_income_2023 - total_expenses_2023
    line_items[12]["value_2024"] = profit_before_exceptional_2024
    line_items[12]["value_2023"] = profit_before_exceptional_2023

    exceptional_items_2024 = line_items[13]["value_2024"]
    exceptional_items_2023 = line_items[13]["value_2023"]
    profit_before_tax_2024 = "{profit_before_tax_2024}"
    profit_before_tax_2023 = "{profit_before_tax_2023}"
    if isinstance(profit_before_exceptional_2024, (int, float)) and isinstance(exceptional_items_2024, (int, float)):
        profit_before_tax_2024 = profit_before_exceptional_2024 - exceptional_items_2024
    if isinstance(profit_before_exceptional_2023, (int, float)) and isinstance(exceptional_items_2023, (int, float)):
        profit_before_tax_2023 = profit_before_exceptional_2023 - exceptional_items_2023
    line_items[14]["value_2024"] = profit_before_tax_2024
    line_items[14]["value_2023"] = profit_before_tax_2023

    current_tax_2024 = line_items[16]["value_2024"]
    current_tax_2023 = line_items[16]["value_2023"]
    deferred_tax_2024 = line_items[17]["value_2024"]
    deferred_tax_2023 = line_items[17]["value_2023"]
    total_tax_2024 = "{total_tax_2024}"
    total_tax_2023 = "{total_tax_2023}"
    if isinstance(current_tax_2024, (int, float)) and isinstance(deferred_tax_2024, (int, float)):
        total_tax_2024 = current_tax_2024 + deferred_tax_2024
    if isinstance(current_tax_2023, (int, float)) and isinstance(deferred_tax_2023, (int, float)):
        total_tax_2023 = current_tax_2023 + deferred_tax_2023
    profit_after_tax_2024 = "{profit_after_tax_2024}"
    profit_after_tax_2023 = "{profit_after_tax_2023}"
    if isinstance(profit_before_tax_2024, (int, float)) and isinstance(total_tax_2024, (int, float)):
        profit_after_tax_2024 = profit_before_tax_2024 - total_tax_2024
    if isinstance(profit_before_tax_2023, (int, float)) and isinstance(total_tax_2023, (int, float)):
        profit_after_tax_2023 = profit_before_tax_2023 - total_tax_2023
    line_items[18]["value_2024"] = profit_after_tax_2024
    line_items[18]["value_2023"] = profit_after_tax_2023

    # Write line items to Excel
    for item in line_items:
        ws.cell(row=row, column=1).value = item["label"]
        ws.cell(row=row, column=2).value = item["note"]
        ws.cell(row=row, column=3).value = item["value_2024"]
        ws.cell(row=row, column=4).value = item["value_2023"]
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align if col > 1 else left_align
            if item["label"].startswith(("I.", "II.", "III.", "IV.", "V.", "VI.", "VII.", "VIII.", "IX.")):
                ws.cell(row=row, column=col).font = bold_font
        row += 1

    # Disclosures
    row += 1
    ws.cell(row=row, column=1).value = "Notes:"
    ws.cell(row=row, column=1).font = bold_font
    row += 1

    # Note 20 disclosure
    note_20_data = load_note_data("20")
    if note_20_data and "notes_and_disclosures" in note_20_data:
        ws.cell(row=row, column=1).value = note_20_data["notes_and_disclosures"][0]
        ws.cell(row=row, column=1).alignment = left_align
        row += 1

    # EPS disclosure
    ws.cell(row=row, column=1).value = "Earnings Per Share (EPS):"
    ws.cell(row=row, column=1).font = bold_font
    row += 1
    ws.cell(row=row, column=1).value = "   Basic EPS"
    ws.cell(row=row, column=3).value = "{basic_eps_2024}"
    ws.cell(row=row, column=4).value = "{basic_eps_2023}"
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=3).alignment = center_align
    ws.cell(row=row, column=4).alignment = center_align
    row += 1
    ws.cell(row=row, column=1).value = "   Diluted EPS"
    ws.cell(row=row, column=3).value = "{diluted_eps_2024}"
    ws.cell(row=row, column=4).value = "{diluted_eps_2023}"
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=3).alignment = center_align
    ws.cell(row=row, column=4).alignment = center_align

    # Save Excel file with error handling
    output_file = "pnl_report.xlsx"
    try:
        wb.save(output_file)
        print(f"P&L report generated and saved to {output_file}")
    except PermissionError:
        print(f"PermissionError: Unable to save to {output_file}. Trying alternative location...")
        fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "pnl_report_fallback.xlsx")
        try:
            wb.save(fallback_file)
            print(f"P&L report saved to alternative location: {fallback_file}")
        except Exception as e:
            print(f"Failed to save P&L report: {str(e)}")
    except Exception as e:
        print(f"Error saving P&L report: {str(e)}")

if __name__ == "__main__":
    generate_pnl_report()