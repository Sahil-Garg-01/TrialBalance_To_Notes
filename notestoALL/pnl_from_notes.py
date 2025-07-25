import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import re

def safe_float_conversion(value):
    """Safely convert value to float, handling various formats."""
    if value is None:
        return 0.0
    
    str_val = str(value).strip()
    
    if not str_val or str_val in ['-', '--', '', 'None']:
        return 0.0
    
    # Remove formatting - Fixed the missing argument
    str_val = str_val.replace(',', '').replace('₹', '').replace('Rs.', '').replace(' ', '')
    
    # Handle parentheses as negative
    is_negative = '(' in str_val and ')' in str_val
    str_val = str_val.replace('(', '').replace(')', '')
    
    try:
        result = float(str_val)
        return -result if is_negative else result
    except (ValueError, TypeError):
        return 0.0

def find_data_columns(ws):
    """Find the data columns for 2024 and 2023."""
    # Look for year headers in first 10 rows
    for row_idx in range(1, 11):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            if "2024" in cell_value:
                col_2024 = col_idx
            elif "2023" in cell_value:
                col_2023 = col_idx
    
    # If not found, use common positions (usually last two data columns)
    if 'col_2024' not in locals():
        col_2024 = ws.max_column - 1 if ws.max_column > 2 else 2
    if 'col_2023' not in locals(): 
        col_2023 = ws.max_column if ws.max_column > 1 else 3
    
    return col_2024, col_2023

def load_and_map_excel_notes(file_path="notestoALL/notes_pnl2.xlsx"):
    """Load and parse notes data from Excel with improved accuracy."""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        notes_data = {}
        
        print(f"Reading Excel file: {file_path}")
        print(f"Sheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Find data columns
        col_2024, col_2023 = find_data_columns(ws)
        print(f"Data columns - 2024: {col_2024}, 2023: {col_2023}")
        
        current_note = None
        current_note_data = None
        
        for row_idx in range(1, ws.max_row + 1):
            # Get row data
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            
            if not row_data or not row_data[0]:
                continue
                
            first_cell = str(row_data[0]).strip()
            
            # Check for note header (16., 17., etc.)
            note_match = re.match(r'^(\d{2})\.?\s*(.+)', first_cell)
            if note_match:
                note_num = note_match.group(1)
                if int(note_num) >= 16:  # Financial notes start from 16
                    # Save previous note if exists
                    if current_note and current_note_data:
                        notes_data[current_note] = current_note_data
                    
                    # Start new note
                    current_note = note_num
                    category_name = note_match.group(2).strip()
                    current_note_data = {
                        "category_name": category_name,
                        "structure": [{
                            "category": category_name,
                            "subcategories": []
                        }],
                        "total_2024": 0.0,
                        "total_2023": 0.0,
                        "total_change": 0.0
                    }
                    print(f"\n📋 Processing Note {note_num}: {category_name}")
                    continue
            
            # Process data rows for current note
            if current_note and current_note_data:
                # Skip header/formatting rows
                skip_keywords = ['in lakhs', 'march 31', 'particulars', 'year ended', 
                               'amount', 'total', 'subtotal', '2024', '2023']
                
                if (len(first_cell) <= 2 or 
                    any(keyword in first_cell.lower() for keyword in skip_keywords)):
                    continue
                
                # Extract values
                value_2024 = 0.0
                value_2023 = 0.0
                
                if col_2024 <= len(row_data):
                    value_2024 = safe_float_conversion(row_data[col_2024 - 1])
                
                if col_2023 <= len(row_data):
                    value_2023 = safe_float_conversion(row_data[col_2023 - 1])
                
                # Add meaningful entries only
                if (first_cell and len(first_cell.strip()) > 2 and 
                    (value_2024 != 0.0 or value_2023 != 0.0 or 
                     any(c.isalpha() for c in first_cell))):
                    
                    subcategory = {
                        "label": first_cell,
                        "value": value_2024,
                        "previous_value": value_2023,
                        "change": value_2024 - value_2023,
                        "change_percent": ((value_2024 - value_2023) / value_2023 * 100) if value_2023 != 0 else 0
                    }
                    
                    current_note_data["structure"][0]["subcategories"].append(subcategory)
                    current_note_data["total_2024"] += value_2024
                    current_note_data["total_2023"] += value_2023
                    
                    print(f"  ✓ {first_cell[:40]:<40} | 2024: {value_2024:>10.2f} | 2023: {value_2023:>10.2f}")
        
        # Save last note
        if current_note and current_note_data:
            current_note_data["total_change"] = current_note_data["total_2024"] - current_note_data["total_2023"]
            notes_data[current_note] = current_note_data
        
        
        
        # Summary - Fixed the variable name
        print("\n" + "="*70)
        print("📈 NOTES SUMMARY")
        print("="*70)
        for note_num in sorted(notes_data.keys(), key=int):
            note = notes_data[note_num]
            items = len(note["structure"][0]["subcategories"])
            print(f"Note {note_num}: {note['category_name'][:35]:<35} | Items: {items:>3} | "
                  f"2024: {note['total_2024']:>10.2f} | 2023: {note['total_2023']:>10.2f}")
        
        return notes_data
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}

def generate_pnl_report(notes_data):
    """Generate comprehensive P&L report matching the exact template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit and Loss Statement"

    # Define styles
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    bold_font = Font(bold=True, size=10)
    
    # Borders
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    top_bottom_border = Border(
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Set column widths to match template
    ws.column_dimensions["A"].width = 45  # Particulars
    ws.column_dimensions["B"].width = 8   # Notes
    ws.column_dimensions["C"].width = 15  # 2024
    ws.column_dimensions["D"].width = 15  # 2023

    row = 1

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Statement of Profit and Loss for the year ended March 31, 2024"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws["A1"].border = top_bottom_border
    row += 1

    # Add empty row
    row += 1

    # Column headers with "In Lakhs" annotation
    ws["A3"] = ""
    ws["B3"] = ""
    ws["C3"] = "In Lakhs"
    ws["D3"] = ""
    ws["C3"].font = normal_font
    ws["C3"].alignment = right_align
    row += 1

    # Main headers
    headers = ["", "Notes", "Year ended March 31, 2024", "Year ended March 31, 2023"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = top_bottom_border
        if col > 2:
            cell.alignment = center_align
        else:
            cell.alignment = left_align
    row += 1

    def add_data_row(description, note_ref, val_2024, val_2023, is_bold=False, is_section_header=False):
        """Add a data row with proper formatting."""
        nonlocal row
        
        # Description
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = description
        cell_a.font = bold_font if (is_bold or is_section_header) else normal_font
        cell_a.alignment = left_align
        if not is_section_header:
            cell_a.border = thin_border
        
        # Note reference
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = note_ref if note_ref else ""
        cell_b.font = normal_font
        cell_b.alignment = center_align
        if not is_section_header:
            cell_b.border = thin_border
        
        # 2024 value
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f"{val_2024:,.2f}" if val_2024 != 0 else ""
        cell_c.font = bold_font if is_bold else normal_font
        cell_c.alignment = right_align
        if not is_section_header:
            cell_c.border = thin_border
        
        # 2023 value
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f"{val_2023:,.2f}" if val_2023 != 0 else ""
        cell_d.font = bold_font if is_bold else normal_font
        cell_d.alignment = right_align
        if not is_section_header:
            cell_d.border = thin_border
        
        row += 1

    # INCOME SECTION
    add_data_row("Income", "", 0, 0, is_section_header=True)
    
    # Revenue from operations
    revenue_2024 = notes_data.get("16", {}).get("total_2024", 0.0)
    revenue_2023 = notes_data.get("16", {}).get("total_2023", 0.0)
    if revenue_2024 == 0 and revenue_2023 == 0:
        print("⚠️ Warning: No data found for Note 16 (Revenue from operations)")
    add_data_row("Revenue from operations (net)", "16", revenue_2024, revenue_2023)
    
    # Other income
    other_income_2024 = notes_data.get("17", {}).get("total_2024", 0.0)
    other_income_2023 = notes_data.get("17", {}).get("total_2023", 0.0)
    if other_income_2024 == 0 and other_income_2023 == 0:
        print("⚠️ Warning: No data found for Note 17 (Other income)")
    add_data_row("Other income", "17", other_income_2024, other_income_2023)
    
    # Total revenue
    total_revenue_2024 = revenue_2024 + other_income_2024
    total_revenue_2023 = revenue_2023 + other_income_2023
    add_data_row("Total revenue (I)", "", total_revenue_2024, total_revenue_2023, is_bold=True)

    # EXPENSES SECTION
    add_data_row("Expenses", "", 0, 0, is_section_header=True)
    
    # Cost of materials consumed
    materials_2024 = notes_data.get("18", {}).get("total_2024", 0.0)
    materials_2023 = notes_data.get("18", {}).get("total_2023", 0.0)
    if materials_2024 == 0 and materials_2023 == 0:
        print("⚠️ Warning: No data found for Note 18 (Cost of materials consumed)")
    add_data_row("Cost of materials consumed", "18", materials_2024, materials_2023)
    
    # Employee benefit expense
    employee_2024 = notes_data.get("19", {}).get("total_2024", 0.0)
    employee_2023 = notes_data.get("19", {}).get("total_2023", 0.0)
    if employee_2024 == 0 and employee_2023 == 0:
        print("⚠️ Warning: No data found for Note 19 (Employee benefit expense)")
    add_data_row("Employee benefit expense", "19", employee_2024, employee_2023)
    
    # Other expenses
    other_exp_2024 = notes_data.get("20", {}).get("total_2024", 0.0)
    other_exp_2023 = notes_data.get("20", {}).get("total_2023", 0.0)
    if other_exp_2024 == 0 and other_exp_2023 == 0:
        print("⚠️ Warning: No data found for Note 20 (Other expenses)")
    add_data_row("Other expenses", "20", other_exp_2024, other_exp_2023)
    
    # Depreciation and amortisation
    depreciation_2024 = notes_data.get("21", {}).get("total_2024", 0.0)
    depreciation_2023 = notes_data.get("21", {}).get("total_2023", 0.0)
    if depreciation_2024 == 0 and depreciation_2023 == 0:
        print("⚠️ Warning: No data found for Note 21 (Depreciation and amortisation expense)")
    add_data_row("Depreciation and amortisation expense", "21", depreciation_2024, depreciation_2023)
    
    # Loss on sale of assets (Note 22)
    loss_sale_2024 = notes_data.get("22", {}).get("total_2024", 0.0)
    loss_sale_2023 = notes_data.get("22", {}).get("total_2023", 0.0)
    if loss_sale_2024 == 0 and loss_sale_2023 == 0 and "22" in notes_data:
        print("⚠️ Warning: No data found for Note 22 (Loss on sale of assets)")
    add_data_row("Loss on sale of assets & investments", "22", loss_sale_2024, loss_sale_2023)
    
    # Finance costs
    finance_2024 = notes_data.get("23", {}).get("total_2024", 0.0)
    finance_2023 = notes_data.get("23", {}).get("total_2023", 0.0)
    if finance_2024 == 0 and finance_2023 == 0:
        print("⚠️ Warning: No data found for Note 23 (Finance costs)")
    add_data_row("Finance costs", "23", finance_2024, finance_2023)
    
    # Total expenses
    total_expenses_2024 = materials_2024 + employee_2024 + other_exp_2024 + depreciation_2024 + loss_sale_2024 + finance_2024
    total_expenses_2023 = materials_2023 + employee_2023 + other_exp_2023 + depreciation_2023 + loss_sale_2023 + finance_2023
    add_data_row("Total Expenses (II)", "", total_expenses_2024, total_expenses_2023, is_bold=True)
    
    # Profit before tax
    profit_before_tax_2024 = total_revenue_2024 - total_expenses_2024
    profit_before_tax_2023 = total_revenue_2023 - total_expenses_2023
    add_data_row("Profit before Tax (I) - (II)", "", profit_before_tax_2024, profit_before_tax_2023, is_bold=True)
    
    # Tax Expense section
    add_data_row("IV. TAX EXPENSE", "", 0, 0, is_section_header=True)
    
    # Current Tax (Placeholder)
    current_tax_2024 = 0.0
    current_tax_2023 = 0.0
    add_data_row("Current Tax", "", current_tax_2024, current_tax_2023)
    
    # Deferred Tax Liability/(Asset) (Placeholder)
    deferred_tax_2024 = 0.0
    deferred_tax_2023 = 0.0
    add_data_row("Deferred Tax Liability/(Asset)", "", deferred_tax_2024, deferred_tax_2023)
    
    # Income Tax relating to Prior Year (Placeholder)
    prior_tax_2024 = 0.0
    prior_tax_2023 = 0.0
    add_data_row("Income Tax relating to Prior Year", "", prior_tax_2024, prior_tax_2023)
    
    # MAT Credit Entitlement/Utilisation (Placeholder)
    mat_credit_2024 = 0.0
    mat_credit_2023 = 0.0
    add_data_row("MAT Credit (Entitlement)/Utilisation", "", mat_credit_2024, mat_credit_2023)
    
    # Total tax
    total_tax_2024 = current_tax_2024 + deferred_tax_2024 + prior_tax_2024 + mat_credit_2024
    total_tax_2023 = current_tax_2023 + deferred_tax_2023 + prior_tax_2023 + mat_credit_2023
    add_data_row("Total Tax Expense (IV)", "", total_tax_2024, total_tax_2023, is_bold=True)
    
    # Profit after Tax
    profit_after_tax_2024 = profit_before_tax_2024 - total_tax_2024
    profit_after_tax_2023 = profit_before_tax_2023 - total_tax_2023
    add_data_row("Profit After Tax (III - IV)", "", profit_after_tax_2024, profit_after_tax_2023, is_bold=True)

    # Earnings per share section
    add_data_row("Earnings per share", "", 0, 0, is_section_header=True)
    
    # Get EPS data from notes
    basic_eps_2024 = 0.0
    basic_eps_2023 = 0.0
    diluted_eps_2024 = 0.0
    diluted_eps_2023 = 0.0
    weighted_shares_2024 = 0.0
    weighted_shares_2023 = 0.0
    
    if "30" in notes_data:
        eps_data = notes_data["30"]["structure"][0]["subcategories"]
        for item in eps_data:
            if "basic" in item["label"].lower():
                basic_eps_2024 = item["value"]
                basic_eps_2023 = item["previous_value"]
            elif "diluted" in item["label"].lower():
                diluted_eps_2024 = item["value"]
                diluted_eps_2023 = item["previous_value"]
            elif "weighted average" in item["label"].lower():
                weighted_shares_2024 = item["value"]
                weighted_shares_2023 = item["previous_value"]
        if basic_eps_2024 == 0 and basic_eps_2023 == 0:
            print("⚠️ Warning: No EPS data found in Note 30")
    
    add_data_row("Basic and diluted", "30", basic_eps_2024, basic_eps_2023)
    add_data_row("Nominal value", "", 10.0, 10.0)  # Assuming ₹10 per share
    
    # Weighted average number of equity shares
    add_data_row("Weighted average number of equity shares", "30", weighted_shares_2024, weighted_shares_2023)

    # Footer notes
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "The accompanying notes are an integral part of the financial statements"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "As per my report of even date                     For and on behalf of the Board of Directors"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 2
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "For M/s Siva Parvathi & Associates                                                    -"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 2
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "ICAI Firm registration number: 020872S"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Chartered Accountants"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align

    # Save the file
    output_folder = "notestoALL"
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, "outputpnl_sheet.xlsx")
    try:
        wb.save(output_file)
        print(f"\n✅ P&L Statement generated: {output_file}")
        
        # Print summary
        print("\n" + "="*60)
        print("📊 FINANCIAL SUMMARY")
        print("="*60)
        print(f"Total Revenue 2024:     ₹{total_revenue_2024:>12,.2f} Lakhs")
        print(f"Total Revenue 2023:     ₹{total_revenue_2023:>12,.2f} Lakhs")
        print(f"Total Expenses 2024:    ₹{total_expenses_2024:>12,.2f} Lakhs")
        print(f"Total Expenses 2023:    ₹{total_expenses_2023:>12,.2f} Lakhs")
        print(f"Profit Before Tax 2024: ₹{profit_before_tax_2024:>12,.2f} Lakhs")
        print(f"Profit Before Tax 2023: ₹{profit_before_tax_2023:>12,.2f} Lakhs")
        print(f"Profit After Tax 2024:  ₹{profit_after_tax_2024:>12,.2f} Lakhs")
        print(f"Profit After Tax 2023:  ₹{profit_after_tax_2023:>12,.2f} Lakhs")
        
        if total_revenue_2023 > 0:
            growth_rate = ((total_revenue_2024 - total_revenue_2023) / total_revenue_2023) * 100
            print(f"Revenue Growth Rate:    {growth_rate:>12.2f}%")
            
        if basic_eps_2024 > 0 or basic_eps_2023 > 0:
            print(f"Basic EPS 2024:         ₹{basic_eps_2024:>12.2f}")
            print(f"Basic EPS 2023:         ₹{basic_eps_2023:>12.2f}")
            print(f"Weighted Shares 2024:   {weighted_shares_2024:>12,.2f} Lakhs")
            print(f"Weighted Shares 2023:   {weighted_shares_2023:>12,.2f} Lakhs")
            
    except PermissionError:
        print(f"❌ Permission Error: Cannot save to {output_file}")
        fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "pnl_statement_fallback.xlsx")
        try:
            wb.save(fallback_file)
            print(f"✅ P&L Statement saved to: {fallback_file}")
        except Exception as e:
            print(f"❌ Failed to save: {str(e)}")
    except Exception as e:
        print(f"❌ Error saving file: {str(e)}")

def main():
    print("🚀 P&L STATEMENT GENERATOR")
    print("=" * 50)
    
    # Step 1: Convert Excel to JSON
    print("\n📋 STEP 1: Converting Excel to JSON")
    notes_data = load_and_map_excel_notes()
    
    if notes_data:
        # Step 2: Generate P&L Statement
        print("\n📊 STEP 2: Generating P&L Statement")
        generate_pnl_report(notes_data)
        print("\n🎉 PROCESS COMPLETED!")
    else:
        print("\n❌ FAILED: No data found")

if __name__ == "__main__":
    main()